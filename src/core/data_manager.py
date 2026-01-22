"""数据管理模块 - 主要数据管理类"""

from typing import List, Dict, Any, Optional, Union, Tuple
import json
import shutil
from pathlib import Path
from datetime import datetime, date
import pandas as pd
import base64
from io import BytesIO
import streamlit as st
import hashlib
import os
import secrets
from openpyxl import Workbook
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.styles import Font, Alignment

from .timeline_manager import TimelineManager
from .models import (
    Project, Experiment, User, BaseModelWithConfig, TimelineInfo,
    RawMaterial, SynthesisRecord, Product, InventoryRecord, MotherLiquor,
    BOM, BOMVersion, ProductionOrder, PasteExperiment, MortarExperiment, ConcreteExperiment,
    ProductInventoryRecord, GoodsReceipt, ShippingOrder, GoodsReceiptItem, ShippingOrderItem, BOMExplosionItem
)
from .enums import (
    ProjectStatus, ExperimentStatus, ExperimentType, UserRole, MotherLiquorSourceType, 
    StockMovementType, BOMStatus, ProductionOrderStatus, IssueStatus, ProductCategory, UnitType, PriorityType, MaterialType, DataCategory,
    ReceiptStatus, ShippingStatus, PermissionAction
)
from config import DATA_FILE, BACKUP_DIR, DEFAULT_UNIT
from .constants import DATE_FORMAT, DATETIME_FORMAT, BACKUP_INTERVAL_SECONDS, DEFAULT_OPERATOR_NAME, DEFAULT_UNIT_KG, WATER_MATERIAL_ALIASES, PRODUCT_NAME_WJSNJ, PRODUCT_NAME_YJSNJ, DEFAULT_BOM_PLAN_QTY
from utils.logger import logger
from utils.unit_helper import convert_quantity, normalize_unit
from utils.file_lock import file_lock

class DataManager:
    """统一数据管理器"""
    
    def __init__(self, data_file_path: Optional[str] = None):
        if data_file_path:
            self.data_file = Path(data_file_path)
            self.backup_dir = Path(data_file_path).parent / "backups"
        else:
            self.data_file = DATA_FILE
            self.backup_dir = BACKUP_DIR
            
        self._data_cache = None
        self._ensure_valid_data_file()
        self._ensure_backup_dir()
        
        # 初始化备份状态
        if "last_backup_time" not in st.session_state:
            st.session_state.last_backup_time = None
    
    def _ensure_valid_data_file(self):
        """确保数据文件存在且格式有效"""
        try:
            # 尝试加载数据，验证文件是否有效
            if self.data_file.exists():
                with open(self.data_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                # 检查数据结构
                if not isinstance(data, dict):
                    raise ValueError("数据格式不正确")
                # 确保有所有必要的数据结构
                self._ensure_data_structure(data)
                return True
        except (json.JSONDecodeError, ValueError, FileNotFoundError):
            # 如果文件无效或不存在，创建初始数据
            print("数据文件无效或不存在，正在创建初始数据...")
            initial_data = self.get_initial_data()
            return self.save_data(initial_data)
        return False
    
    def _ensure_data_structure(self, data):
        """确保数据结构完整"""
        required_keys = [
            DataCategory.PROJECTS.value, DataCategory.EXPERIMENTS.value, DataCategory.PERFORMANCE_DATA.value,
            DataCategory.RAW_MATERIALS.value, DataCategory.SYNTHESIS_RECORDS.value, DataCategory.PRODUCTS.value,
            DataCategory.PASTE_EXPERIMENTS.value, DataCategory.MORTAR_EXPERIMENTS.value, DataCategory.CONCRETE_EXPERIMENTS.value,
            DataCategory.MOTHER_LIQUORS.value,
            # SAP/BOM Modules
            DataCategory.INVENTORY_RECORDS.value, DataCategory.BOMS.value, DataCategory.BOM_VERSIONS.value, 
            DataCategory.PRODUCTION_ORDERS.value, DataCategory.MATERIAL_ISSUES.value, DataCategory.GOODS_RECEIPTS.value, DataCategory.SHIPPING_ORDERS.value,
            # 用户与权限
            DataCategory.USERS.value, DataCategory.AUDIT_LOGS.value
        ]
        
        for key in required_keys:
            if key not in data:
                if key == DataCategory.PERFORMANCE_DATA.value:
                    data[key] = {"synthesis": [], "paste": [], "mortar": [], "concrete": []}
                else:
                    data[key] = []
        
        return data
    
    def _ensure_backup_dir(self):
        """确保备份目录存在"""
        self.backup_dir.mkdir(parents=True, exist_ok=True)

    def _get_next_id(self, items: List[Dict[str, Any]]) -> int:
        """Safely calculate the next integer ID, ignoring non-integer IDs."""
        ids = []
        for item in items:
            val = item.get("id")
            if isinstance(val, int):
                ids.append(val)
            elif isinstance(val, str) and val.isdigit():
                try:
                    ids.append(int(val))
                except:
                    pass
        return max(ids, default=0) + 1
    
    def load_data(self):
        """从JSON文件加载所有数据"""
        try:
            if self._data_cache is not None:
                return self._ensure_data_structure(self._data_cache)
            if self.data_file.exists():
                with open(self.data_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                data = self._ensure_data_structure(data)
                self._data_cache = data
                return data
            initial = self.get_initial_data()
            self._data_cache = initial
            return initial
        except Exception as e:
            st.error(f"读取数据失败: {e}")
            initial = self.get_initial_data()
            self._data_cache = initial
            return initial
    
    def save_data(self, data):
        """保存数据到JSON文件，并创建备份"""
        try:
            self.data_file.parent.mkdir(parents=True, exist_ok=True)
            
            # 使用文件锁防止并发写入冲突
            with file_lock(self.data_file, timeout=10):
                # 1. 强制备份 (Atomic Pre-write Backup)
                self.create_backup(force=True)
                
                # 2. 原子写入 (Write to temp -> Flush -> Rename)
                temp_file = self.data_file.with_suffix('.tmp')
                
                with open(temp_file, 'w', encoding='utf-8') as f:
                    json.dump(data, f, ensure_ascii=False, indent=4)
                    f.flush()
                    os.fsync(f.fileno()) # 确保数据落盘
                
                # 原子替换 (Windows上如果目标存在可能会有问题，但在context manager保护下相对安全，且os.replace在py3中尽可能原子)
                if temp_file.exists():
                    try:
                        os.replace(temp_file, self.data_file)
                    except OSError:
                         # Windows fallback: remove target then rename
                        if os.name == 'nt' and self.data_file.exists():
                            try:
                                os.remove(self.data_file)
                                os.rename(temp_file, self.data_file)
                            except Exception as e:
                                logger.error(f"Windows atomic write fallback failed: {e}")
                                raise
                        else:
                            raise
            
            self._data_cache = self._ensure_data_structure(data)
            # 记录最后备份时间 (虽然create_backup(force=True)已经做了，但为了兼容check_and_create_auto_backup逻辑)
            st.session_state.last_backup_time = datetime.now()
            
            return True
        except Exception as e:
            st.error(f"保存数据失败: {e}")
            logger.error(f"Save data failed: {e}")
            return False
    
    def check_and_create_auto_backup(self):
        """
        检查并创建自动备份
        (注：save_data 现已强制备份，此方法主要用于非写入触发的定时检查，或者保持兼容性)
        """
        try:
            # 获取当前时间
            now = datetime.now()
            
            # 检查上次备份时间
            last_backup = st.session_state.get("last_backup_time")
            
            should_backup = False
            if last_backup is None:
                should_backup = True
            elif isinstance(last_backup, datetime):
                # 如果上次备份超过设定间隔
                if (now - last_backup).total_seconds() > BACKUP_INTERVAL_SECONDS:
                    should_backup = True
            else:
                should_backup = True
                
            if should_backup:
                self.create_backup()
                st.session_state.last_backup_time = now
        except Exception as e:
            print(f"检查备份失败: {e}")
    
    def create_backup(self, force=False):
        """
        创建数据备份
        Args:
            force: 是否强制备份（忽略间隔检查，通常用于写入前）
        """
        try:
            if self.data_file.exists():
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f") # 增加微秒以防快速连续写入
                backup_file = self.backup_dir / f"data_backup_{timestamp}.json"
                
                # 复制文件
                shutil.copy2(self.data_file, backup_file)
                
                # 清理旧的备份文件（保留最近50个，防止写入频繁导致磁盘满）
                self._cleanup_old_backups(max_backups=50)
                
                return True
        except Exception as e:
            print(f"创建备份失败: {e}")
            logger.error(f"Create backup failed: {e}")
            return False
    
    def _cleanup_old_backups(self, max_backups=30):
        """清理旧的备份文件"""
        try:
            backup_files = list(self.backup_dir.glob("data_backup_*.json"))
            
            if len(backup_files) > max_backups:
                # 按修改时间排序，删除最旧的文件
                backup_files.sort(key=lambda x: x.stat().st_mtime)
                files_to_delete = backup_files[:-max_backups]
                
                for file in files_to_delete:
                    file.unlink()
                    
        except Exception as e:
            print(f"清理备份文件失败: {e}")


    
    # -------------------- 获取初始数据 --------------------
    def get_initial_data(self) -> Dict[str, Any]:
        """返回初始数据结构"""
        return {
            DataCategory.PROJECTS.value: [
                {
                    "id": 1,
                    "name": "PC-001合成优化",
                    "leader": "张三",
                    "start_date": "2024-01-01",
                    "end_date": "2024-03-01",
                    "status": ProjectStatus.IN_PROGRESS.value,
                    "progress": 75,
                    "description": "优化聚羧酸合成工艺参数"
                },
                {
                    "id": 2,
                    "name": "PC-002性能测试",
                    "leader": "李四",
                    "start_date": "2024-01-10",
                    "end_date": "2024-02-15",
                    "status": ProjectStatus.COMPLETED.value,
                    "progress": 100,
                    "description": "测试不同配方性能"
                }
            ],
            DataCategory.EXPERIMENTS.value: [
                {
                    "id": 1,
                    "name": "PC-001-合成实验1",
                    "type": ExperimentType.SYNTHESIS.value,
                    "project_id": 1,
                    "planned_date": "2024-01-20",
                    "actual_date": "2024-01-20",
                    "priority": PriorityType.HIGH.value,
                    "status": ExperimentStatus.COMPLETED.value,
                    "description": "第一轮合成实验"
                }
            ],
            DataCategory.PERFORMANCE_DATA.value: {
                "synthesis": [
                    {
                        "id": 1,
                        "batch": "PC-001",
                        "water_reduction": 18.5,
                        "solid_content": 40,
                        "slump_flow": 650,
                        "test_date": "2024-01-10",
                        "sample_id": "PC-001-20240110"
                    }
                ],
                "paste": [],
                "mortar": [],
                "concrete": []
            },
            DataCategory.RAW_MATERIALS.value: [],
            DataCategory.SYNTHESIS_RECORDS.value: [],
            DataCategory.PRODUCTS.value: [],
            DataCategory.PASTE_EXPERIMENTS.value: [],
            DataCategory.MORTAR_EXPERIMENTS.value: [],
            DataCategory.CONCRETE_EXPERIMENTS.value: [],
            DataCategory.MOTHER_LIQUORS.value: [],
            DataCategory.INVENTORY_RECORDS.value: [],
            DataCategory.BOMS.value: [],
            DataCategory.BOM_VERSIONS.value: [],
            DataCategory.PRODUCTION_ORDERS.value: [],
            DataCategory.MATERIAL_ISSUES.value: [],
            DataCategory.GOODS_RECEIPTS.value: [],
            DataCategory.SHIPPING_ORDERS.value: [],
            DataCategory.USERS.value: [],
            DataCategory.AUDIT_LOGS.value: []
        }
    
    # -------------------- 项目CRUD操作 --------------------
    def get_all_projects(self) -> List[Dict[str, Any]]:
        """获取所有项目"""
        data = self.load_data()
        return data.get(DataCategory.PROJECTS.value, [])

    # -------------------- 用户与权限管理 --------------------
    def _hash_password(self, password: str, salt: str) -> str:
        value = f"{salt}:{password}".encode("utf-8")
        return hashlib.sha256(value).hexdigest()
    
    def get_all_users(self) -> List[Dict[str, Any]]:
        data = self.load_data()
        return data.get(DataCategory.USERS.value, [])

    def get_user_by_username(self, username):
        username = str(username or "").strip().lower()
        users = self.get_all_users()
        for u in users:
            if str(u.get("username", "")).strip().lower() == username:
                return u
        return None

    def create_user(self, username, password, role=UserRole.USER.value):
        username = str(username or "").strip()
        if not username or not password:
            return False, "用户名和密码不能为空"
        if role != UserRole.ADMIN.value:
            parts = username.split()
            if len(parts) != 2:
                return False, "用户名格式应为“姓名 手机号”"
            name, mobile = parts[0], parts[1]
            mobile_str = str(mobile)
            if not mobile_str.isdigit() or len(mobile_str) != 11 or not mobile_str.startswith(("1",)):
                return False, "手机号格式不正确，请输入11位手机号码"
        if self.get_user_by_username(username):
            return False, "用户名已存在"
        data = self.load_data()
        users = data.get(DataCategory.USERS.value, [])
        new_id = self._get_next_id(users)
        salt = secrets.token_hex(16)
        pwd_hash = self._hash_password(password, salt)
        role_norm = role if role in [UserRole.ADMIN.value, UserRole.USER.value] else UserRole.USER.value
        users.append({
            "id": new_id,
            "username": username,
            "password_hash": pwd_hash,
            "salt": salt,
            "role": role_norm,
            "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "active": True
        })
        data[DataCategory.USERS.value] = users
        if self.save_data(data):
            return True, "注册成功"
        return False, "保存失败"

    def authenticate_user(self, username, password):
        user = self.get_user_by_username(username)
        if not user or not user.get("active", True):
            return False, None
        salt = user.get("salt", "")
        expected = user.get("password_hash", "")
        calc = self._hash_password(password, salt)
        if secrets.compare_digest(calc, expected):
            return True, {"id": user.get("id"), "username": user.get("username"), "role": user.get("role", UserRole.USER.value)}
        return False, None

    def change_user_password(self, user_id, old_password, new_password):
        data = self.load_data()
        users = data.get(DataCategory.USERS.value, [])
        target_index = -1
        for i, u in enumerate(users):
            if u.get("id") == user_id:
                target_index = i
                break
        if target_index == -1:
            return False, "用户不存在"
        user = users[target_index]
        salt = user.get("salt", "")
        expected = user.get("password_hash", "")
        if not salt or not expected:
            return False, "当前密码校验失败"
        calc = self._hash_password(old_password, salt)
        if not secrets.compare_digest(calc, expected):
            return False, "当前密码错误"
        new_salt = secrets.token_hex(16)
        new_hash = self._hash_password(new_password, new_salt)
        user["salt"] = new_salt
        user["password_hash"] = new_hash
        users[target_index] = user
        data[DataCategory.USERS.value] = users
        if self.save_data(data):
            return True, "登录密码已更新"
        return False, "保存失败"

    def ensure_default_admin(self):
        users = self.get_all_users()
        has_admin = any(u.get("role") == UserRole.ADMIN.value for u in users)
        if has_admin:
            return
        default_pwd = os.environ.get("APP_ADMIN_PASSWORD") or "admin"
        ok, msg = self.create_user("admin", default_pwd, role=UserRole.ADMIN.value)
        if ok:
            logger.info("Default admin user created.")

    def update_user(self, user_id, fields):
        data = self.load_data()
        users = data.get(DataCategory.USERS.value, [])
        updated = False
        for i, u in enumerate(users):
            if u.get("id") == user_id:
                users[i].update(fields)
                updated = True
                break
        if updated:
            data[DataCategory.USERS.value] = users
            return self.save_data(data)
        return False

    def get_admin_users(self):
        users = self.get_all_users()
        return [u for u in users if u.get("role") == UserRole.ADMIN.value and u.get("active", True)]
    
    def add_audit_log(self, user, action: str, detail: str):
        data = self.load_data()
        logs = data.get(DataCategory.AUDIT_LOGS.value, [])
        new_id = self._get_next_id(logs)
        entry = {
            "id": new_id,
            "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "user_id": user.get("id") if user else None,
            "username": user.get("username") if user else "系统",
            "role": user.get("role", "") if user else "",
            "action": str(action),
            "detail": str(detail),
        }
        logs.append(entry)
        if len(logs) > 1000:
            logs = logs[-1000:]
        data[DataCategory.AUDIT_LOGS.value] = logs
        return self.save_data(data)
    
    def get_audit_logs(self):
        data = self.load_data()
        return data.get(DataCategory.AUDIT_LOGS.value, [])
    
    def get_next_project_id(self):
        """获取下一个可用的项目ID"""
        projects = self.get_all_projects()
        if not projects:
            return 1
        return self._get_next_id(projects)
    
    def get_project(self, project_id: int) -> Optional[Dict[str, Any]]:
        """根据ID获取单个项目"""
        projects = self.get_all_projects()
        for project in projects:
            if project.get("id") == project_id:
                return project
        return None
    
    def add_project(self, project_data: Union[Dict[str, Any], Project]) -> bool:
        """添加新项目"""
        data = self.load_data()
        projects = data.get(DataCategory.PROJECTS.value, [])
        
        # 生成新ID
        new_id = self._get_next_id(projects)
        
        if isinstance(project_data, Project):
            project_data.id = new_id
            # mode='json' 会将 date 对象转换为 ISO 格式字符串
            p_dict = project_data.model_dump(mode='json')
        else:
            project_data["id"] = new_id
            
            # 确保日期是字符串格式
            for date_field in ["start_date", "end_date"]:
                if date_field in project_data and hasattr(project_data[date_field], 'strftime'):
                    project_data[date_field] = project_data[date_field].strftime("%Y-%m-%d")
            
            # 尝试使用模型验证（不强制阻断，但建议）
            try:
                Project(**project_data)
            except Exception as e:
                logger.warning(f"Project data validation warning: {e}")
            
            p_dict = project_data
        
        projects.append(p_dict)
        data[DataCategory.PROJECTS.value] = projects
        success = self.save_data(data)
        return success
    
    def update_project(self, project_id: int, updated_fields: Dict[str, Any]) -> bool:
        """更新项目信息"""
        data = self.load_data()
        projects = data.get(DataCategory.PROJECTS.value, [])
        
        updated = False
        for i, project in enumerate(projects):
            if project.get("id") == project_id:
                # 更新字段
                projects[i].update(updated_fields)
                updated = True
                break
        
        if updated:
            data[DataCategory.PROJECTS.value] = projects
            return self.save_data(data)
        return False
    
    def delete_project(self, project_id: int) -> bool:
        """根据ID删除项目"""
        try:
            data = self.load_data()
            projects = data.get(DataCategory.PROJECTS.value, [])
            
            # 记录删除前的数量
            original_count = len(projects)
            
            # 过滤掉要删除的项目
            new_projects = [p for p in projects if p.get("id") != project_id]
            
            # 检查是否真的删除了项目
            if len(new_projects) < original_count:
                data[DataCategory.PROJECTS.value] = new_projects
                success = self.save_data(data)
                if success:
                    print(f"成功删除项目 ID: {project_id}")
                    return True
                else:
                    print(f"保存数据失败，项目 ID: {project_id} 删除未生效")
                    return False
            else:
                print(f"未找到项目 ID: {project_id}")
                return False
        except Exception as e:
            print(f"删除项目时出错: {e}")
            return False
    
    # -------------------- 项目时间线相关方法 --------------------
    def get_project_timeline(self, project_id: int) -> Optional[TimelineInfo]:
        """
        获取项目时间线信息（使用TimelineManager）
        参数:
            project_id: 项目ID
        返回:
            时间线信息对象，如果项目不存在返回None
        """
        # 1. 获取项目数据
        project_data = self.get_project(project_id)
        if not project_data:
            logger.warning(f"警告: 未找到项目ID {project_id}")
            return None
        
        # 2. 使用TimelineManager进行计算
        return TimelineManager.calculate_timeline(project_data)
    
    # -------------------- 实验CRUD操作 --------------------
    def get_all_experiments(self) -> List[Dict[str, Any]]:
        """获取所有实验"""
        data = self.load_data()
        return data.get(DataCategory.EXPERIMENTS.value, [])
    
    def add_experiment(self, experiment_data: Union[Dict[str, Any], Experiment]) -> bool:
        """添加新实验"""
        data = self.load_data()
        experiments = data.get(DataCategory.EXPERIMENTS.value, [])
        
        # 生成新ID
        new_id = self._get_next_id(experiments)
        
        if isinstance(experiment_data, Experiment):
            experiment_data.id = new_id
            exp_dict = experiment_data.model_dump(mode='json')
        else:
            experiment_data["id"] = new_id
            
            # 确保日期是字符串格式
            for date_field in ["planned_date", "actual_date"]:
                if date_field in experiment_data and experiment_data[date_field]:
                    if hasattr(experiment_data[date_field], 'strftime'):
                        experiment_data[date_field] = experiment_data[date_field].strftime("%Y-%m-%d")
            
            # 尝试使用模型验证
            try:
                Experiment(**experiment_data)
            except Exception as e:
                logger.warning(f"Experiment data validation warning: {e}")
                
            exp_dict = experiment_data
        
        experiments.append(exp_dict)
        data[DataCategory.EXPERIMENTS.value] = experiments
        return self.save_data(data)
    
    def update_experiment(self, experiment_id: int, updated_fields: Dict[str, Any]) -> bool:
        """更新实验信息"""
        data = self.load_data()
        experiments = data.get(DataCategory.EXPERIMENTS.value, [])
        
        updated = False
        for i, exp in enumerate(experiments):
            if exp.get("id") == experiment_id:
                # 更新字段
                experiments[i].update(updated_fields)
                updated = True
                break
        
        if updated:
            data[DataCategory.EXPERIMENTS.value] = experiments
            return self.save_data(data)
        return False
    
    def delete_experiment(self, experiment_id: int) -> bool:
        """根据ID删除实验"""
        data = self.load_data()
        experiments = data.get(DataCategory.EXPERIMENTS.value, [])
        
        new_experiments = [e for e in experiments if e.get("id") != experiment_id]
        
        if len(new_experiments) < len(experiments):
            data[DataCategory.EXPERIMENTS.value] = new_experiments
            return self.save_data(data)
        return False
    
    # -------------------- 数据记录模块CRUD操作 --------------------
    # 原材料管理
    def get_all_raw_materials(self):
        """获取所有原材料"""
        data = self.load_data()
        return data.get(DataCategory.RAW_MATERIALS.value, [])
    
    def add_raw_material(self, material_data: Union[Dict[str, Any], RawMaterial]) -> tuple[bool, str]:
        """添加新原材料"""
        data = self.load_data()
        materials = data.get(DataCategory.RAW_MATERIALS.value, [])
        
        # Determine name to check for duplicates
        if isinstance(material_data, RawMaterial):
            name = material_data.name
        else:
            name = material_data.get("name")
        
        # 检查名称是否重复
        if name:
            for m in materials:
                if m.get("name") == name:
                    return False, f"原材料名称 '{name}' 已存在"
        
        # 生成新ID
        new_id = self._get_next_id(materials)
        
        if isinstance(material_data, RawMaterial):
            material_data.id = new_id
            if not material_data.created_date:
                material_data.created_date = datetime.now().strftime("%Y-%m-%d")
            mat_dict = material_data.model_dump(mode='json')
        else:
            material_data["id"] = new_id
            material_data["created_date"] = datetime.now().strftime("%Y-%m-%d")
            
            # Try validation
            try:
                RawMaterial(**material_data)
            except Exception as e:
                logger.warning(f"RawMaterial validation warning: {e}")
                
            mat_dict = material_data
        
        materials.append(mat_dict)
        data[DataCategory.RAW_MATERIALS.value] = materials
        if self.save_data(data):
            return True, "添加成功"
        return False, "保存失败"
    
    def update_raw_material(self, material_id: int, updated_fields: Dict[str, Any]) -> tuple[bool, str]:
        """更新原材料信息"""
        try:
            data = self.load_data()
            materials = data.get(DataCategory.RAW_MATERIALS.value, [])
            
            # 验证名称唯一性
            new_name = updated_fields.get("name")
            if new_name:
                # 检查是否有其他原材料使用了该名称
                for m in materials:
                    if m.get("id") != material_id and m.get("name") == new_name:
                        msg = f"原材料名称 '{new_name}' 已存在"
                        logger.warning(f"Update failed: {msg}")
                        return False, msg
            
            old_name = None
            updated = False
            for i, material in enumerate(materials):
                if material.get("id") == material_id:
                    old_name = material.get("name")
                    # 更新字段
                    materials[i].update(updated_fields)
                    updated = True
                    break
            
            if updated:
                # 如果名称发生了变化，更新所有引用该名称的记录
                if new_name and old_name and new_name != old_name:
                    logger.info(f"Raw material renamed from '{old_name}' to '{new_name}'. Updating references...")
                    self._update_raw_material_references(data, old_name, new_name)
                
                data[DataCategory.RAW_MATERIALS.value] = materials
                if self.save_data(data):
                    logger.info(f"Raw material {material_id} updated successfully.")
                    return True, "更新成功"
                else:
                    msg = "保存数据失败"
                    logger.error(f"{msg} after updating raw material {material_id}.")
                    return False, msg
            
            msg = f"原材料 ID {material_id} 未找到"
            logger.warning(msg)
            return False, msg
        except Exception as e:
            msg = f"系统错误: {e}"
            logger.error(f"Error updating raw material {material_id}: {e}")
            return False, msg
    
    def _update_raw_material_references(self, data: Dict[str, Any], old_name: str, new_name: str) -> None:
        """更新所有引用旧原材料名称的记录"""
        try:
            count = 0
            # 1. 更新合成实验记录
            records = data.get(DataCategory.SYNTHESIS_RECORDS.value, [])
            for record in records:
                modified = False
                # 检查所有物料列表
                for key in ["reactor_materials", "a_materials", "b_materials"]:
                    items = record.get(key, [])
                    if items:
                        for item in items:
                            if item.get("material_name") == old_name:
                                item["material_name"] = new_name
                                modified = True
                                count += 1
                
                if modified:
                    record["last_modified"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            # 2. 更新成品减水剂配方
            products = data.get(DataCategory.PRODUCTS.value, [])
            for product in products:
                modified = False
                # Check "ingredients" (standard) and "composition" (legacy)
                for key in ["ingredients", "composition"]:
                    items = product.get(key, [])
                    if items:
                        for item in items:
                            if item.get("name") == old_name:
                                item["name"] = new_name
                                modified = True
                                count += 1
                
                if modified:
                    product["last_modified"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            if count > 0:
                logger.info(f"Updated {count} references from '{old_name}' to '{new_name}'")
            else:
                logger.info(f"No references found for '{old_name}'")
                
        except Exception as e:
            logger.error(f"Error updating references: {e}")
    
    def delete_raw_material(self, material_id: int) -> tuple[bool, str]:
        """删除原材料"""
        try:
            data = self.load_data()
            materials = data.get(DataCategory.RAW_MATERIALS.value, [])
            
            # 获取要删除的原材料名称
            material_to_delete = next((m for m in materials if m.get("id") == material_id), None)
            if not material_to_delete:
                logger.warning(f"Delete failed: Raw material ID {material_id} not found.")
                return False, "原材料不存在"
            
            material_name = material_to_delete.get("name")
            
            # 检查引用
            # 1. 检查合成实验记录
            records = data.get(DataCategory.SYNTHESIS_RECORDS.value, [])
            for record in records:
                for key in ["reactor_materials", "a_materials", "b_materials"]:
                    items = record.get(key, [])
                    for item in items:
                        if item.get("material_name") == material_name:
                            msg = f"无法删除：该原材料在合成实验配方 {record.get('formula_id', '未知')} 中被使用"
                            logger.warning(f"Delete blocked: {msg}")
                            return False, msg
            
            # 2. 检查成品减水剂
            products = data.get(DataCategory.PRODUCTS.value, [])
            for product in products:
                # 检查 ingredients
                ingredients = product.get("ingredients", [])
                for item in ingredients:
                    if item.get("name") == material_name:
                        msg = f"无法删除：该原材料在成品 {product.get('product_name', '未知')} 中被使用"
                        logger.warning(f"Delete blocked: {msg}")
                        return False, msg
                
                # 检查 composition (兼容旧数据)
                composition = product.get("composition", [])
                for item in composition:
                    if item.get("name") == material_name:
                        msg = f"无法删除：该原材料在成品 {product.get('product_name', '未知')} 中被使用"
                        logger.warning(f"Delete blocked: {msg}")
                        return False, msg

            new_materials = [m for m in materials if m.get("id") != material_id]
            
            if len(new_materials) < len(materials):
                data[DataCategory.RAW_MATERIALS.value] = new_materials
                if self.save_data(data):
                    logger.info(f"Raw material {material_id} ({material_name}) deleted successfully.")
                    return True, "删除成功"
                else:
                    logger.error(f"Failed to save data after deleting raw material {material_id}.")
                    return False, "保存数据失败"
            
            logger.warning(f"Delete failed: Raw material {material_id} deletion not effective.")
            return False, "删除操作未生效"
            
        except Exception as e:
            logger.error(f"Error deleting raw material {material_id}: {e}")
            return False, f"系统错误: {str(e)}"

    # -------------------- 库存管理 (SAP Lite) --------------------
    def add_inventory_record(self, record_data: Union[Dict[str, Any], InventoryRecord], update_master_stock: bool = True) -> tuple[bool, str]:
        """添加库存变动记录"""
        try:
            data = self.load_data()
            
            # Convert model to dict if necessary
            if isinstance(record_data, InventoryRecord):
                 rec_dict = record_data.model_dump(mode='json')
            else:
                 rec_dict = record_data

            # 1. Update Stock in Raw Material
            materials = data.get(DataCategory.RAW_MATERIALS.value, [])
            material_found = False
            current_stock = 0.0
            
            material_id = rec_dict.get("material_id")
            if isinstance(material_id, str) and material_id.isdigit():
                material_id = int(material_id)
            
            for i, m in enumerate(materials):
                if m.get("id") == material_id:
                    # Check if material is "Water" (no stock tracking)
                    mat_name = m.get("name", "").strip()
                    water_names = ["水", "自来水", "纯水", "去离子水", "工业用水", "生产用水"]
                    is_water = mat_name in water_names
                    
                    if not is_water:
                        current_stock = float(m.get("stock_quantity", 0.0))
                        
                        if update_master_stock:
                            change = float(rec_dict.get("quantity", 0.0))
                            
                            # Use Enum or string
                            rec_type = rec_dict.get("type")
                            if rec_type == "out" or rec_type == StockMovementType.OUT:
                                current_stock -= change
                            else:
                                current_stock += change
                            
                            materials[i]["stock_quantity"] = current_stock
                            materials[i]["last_stock_update"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    
                    material_found = True
                    break
            
            if not material_found:
                return False, "原材料不存在"

            # 2. Add Record to Ledger
            records = data.get(DataCategory.INVENTORY_RECORDS.value, [])
            new_id = self._get_next_id(records)
            
            if isinstance(record_data, InventoryRecord):
                record_data.id = new_id
                record_data.snapshot_stock = current_stock
                if not record_data.created_at:
                    record_data.created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                final_record = record_data.model_dump(mode='json')
            else:
                rec_dict["id"] = new_id
                rec_dict["created_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                rec_dict["snapshot_stock"] = current_stock
                
                # Validation attempt
                try:
                    InventoryRecord(**rec_dict)
                except Exception as e:
                    logger.warning(f"InventoryRecord validation warning: {e}")
                
                final_record = rec_dict
            
            records.append(final_record)
            
            data[DataCategory.RAW_MATERIALS.value] = materials
            data[DataCategory.INVENTORY_RECORDS.value] = records
            
            if self.save_data(data):
                return True, "库存更新成功"
            return False, "保存失败"
        except Exception as e:
            logger.error(f"Error adding inventory record: {e}")
            return False, f"系统错误: {str(e)}"

    def get_inventory_records(self, material_id=None):
        """获取库存记录"""
        data = self.load_data()
        records = data.get(DataCategory.INVENTORY_RECORDS.value, [])
        if material_id:
            if isinstance(material_id, str) and material_id.isdigit():
                material_id = int(material_id)
            return [r for r in records if r.get("material_id") == material_id]
        return records
    
    # -------------------- 母液管理CRUD操作 --------------------
    def get_all_mother_liquors(self):
        """获取所有母液"""
        data = self.load_data()
        return data.get(DataCategory.MOTHER_LIQUORS.value, [])
    
    def get_mother_liquor(self, ml_id):
        """根据ID获取单个母液"""
        mls = self.get_all_mother_liquors()
        for ml in mls:
            if ml.get("id") == ml_id:
                return ml
        return None

    def add_mother_liquor(self, ml_data: Union[Dict[str, Any], MotherLiquor]) -> bool:
        """添加新母液"""
        data = self.load_data()
        mls = data.get(DataCategory.MOTHER_LIQUORS.value, [])
        
        # 生成新ID
        new_id = self._get_next_id(mls)
        
        if isinstance(ml_data, MotherLiquor):
            ml_data.id = new_id
            if not ml_data.created_at:
                ml_data.created_at = datetime.now().strftime("%Y-%m-%d %H:%M")
            final_ml = ml_data.model_dump(mode='json')
        else:
            ml_data["id"] = new_id
            
            # Validation attempt
            try:
                MotherLiquor(**ml_data)
            except Exception as e:
                logger.warning(f"MotherLiquor validation warning: {e}")
                
            final_ml = ml_data
        
        mls.append(final_ml)
        data[DataCategory.MOTHER_LIQUORS.value] = mls
        return self.save_data(data)

    def update_mother_liquor(self, ml_id: int, updated_fields: Dict[str, Any]) -> bool:
        """更新母液信息"""
        data = self.load_data()
        mls = data.get(DataCategory.MOTHER_LIQUORS.value, [])
        
        updated = False
        for i, ml in enumerate(mls):
            if ml.get("id") == ml_id:
                # 更新字段
                mls[i].update(updated_fields)
                mls[i]["last_modified"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                updated = True
                break
        
        if updated:
            data[DataCategory.MOTHER_LIQUORS.value] = mls
            return self.save_data(data)
        return False

    def delete_mother_liquor(self, ml_id: int) -> bool:
        """删除母液"""
        data = self.load_data()
        mls = data.get(DataCategory.MOTHER_LIQUORS.value, [])
        
        new_mls = [m for m in mls if m.get("id") != ml_id]
        
        if len(new_mls) < len(mls):
            data[DataCategory.MOTHER_LIQUORS.value] = new_mls
            return self.save_data(data)
        return False
    
    # -------------------- 合成实验记录CRUD操作 --------------------
    def get_synthesis_experiment(self, exp_id: int) -> Optional[Dict[str, Any]]:
        """获取单个合成实验记录"""
        data = self.load_data()
        records = data.get(DataCategory.SYNTHESIS_RECORDS.value, [])
        for record in records:
            if record.get("id") == exp_id:
                return record
        return None

    def get_all_synthesis_experiments(self) -> List[Dict[str, Any]]:
        """获取所有合成实验（别名方法，兼容调用）"""
        return self.get_all_synthesis_records()
        
    def get_all_synthesis_records(self) -> List[Dict[str, Any]]:
        """获取所有合成实验记录"""
        data = self.load_data()
        return data.get(DataCategory.SYNTHESIS_RECORDS.value, [])
    
    def add_synthesis_record(self, record_data: Union[Dict[str, Any], SynthesisRecord]) -> bool:
        """添加新合成实验记录"""
        data = self.load_data()
        records = data.get(DataCategory.SYNTHESIS_RECORDS.value, [])
        
        # 生成新ID
        new_id = self._get_next_id(records)
        
        if isinstance(record_data, SynthesisRecord):
            record_data.id = new_id
            now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            if not record_data.created_at:
                record_data.created_at = now_str
            record_data.last_modified = now_str
            rec_dict = record_data.model_dump(mode='json')
        else:
            record_data["id"] = new_id
            now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            record_data["created_at"] = now_str
            record_data["last_modified"] = now_str
            
            try:
                SynthesisRecord(**record_data)
            except Exception as e:
                logger.warning(f"SynthesisRecord validation warning: {e}")
                
            rec_dict = record_data
        
        records.append(rec_dict)
        data[DataCategory.SYNTHESIS_RECORDS.value] = records
        return self.save_data(data)
    
    def delete_synthesis_record(self, record_id: int) -> bool:
        """删除合成实验记录"""
        data = self.load_data()
        records = data.get(DataCategory.SYNTHESIS_RECORDS.value, [])
        
        new_records = [r for r in records if r.get("id") != record_id]
        
        if len(new_records) < len(records):
            data[DataCategory.SYNTHESIS_RECORDS.value] = new_records
            return self.save_data(data)
        return False
    
    # 成品减水剂
    def get_all_products(self) -> List[Dict[str, Any]]:
        """获取所有成品减水剂"""
        data = self.load_data()
        return data.get(DataCategory.PRODUCTS.value, [])

    def check_product_name_exists(self, name: str, exclude_id: Optional[int] = None) -> bool:
        """检查成品名称是否已存在（不区分大小写）"""
        products = self.get_all_products()
        name = name.strip().lower()
        for product in products:
            if exclude_id is not None and product.get("id") == exclude_id:
                continue
            if product.get("product_name", "").strip().lower() == name:
                return True
        return False
    
    def add_product(self, product_data: Union[Dict[str, Any], Product]) -> bool:
        """添加新成品减水剂"""
        
        if isinstance(product_data, Product):
            p_name = product_data.product_name
        else:
            p_name = product_data.get("product_name", "")
            
        # 检查名称唯一性
        if self.check_product_name_exists(p_name):
             raise ValueError(f"成品名称 '{p_name}' 已存在")

        data = self.load_data()
        products = data.get(DataCategory.PRODUCTS.value, [])
        
        # 生成新ID
        new_id = self._get_next_id(products)
        
        if isinstance(product_data, Product):
            product_data.id = new_id
            product_data.created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            prod_dict = product_data.model_dump(mode='json')
        else:
            product_data["id"] = new_id
            product_data["created_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            try:
                Product(**product_data)
            except Exception as e:
                logger.warning(f"Product validation warning: {e}")
                
            prod_dict = product_data
        
        products.append(prod_dict)
        data[DataCategory.PRODUCTS.value] = products
        return self.save_data(data)
    
    def delete_product(self, product_id: int) -> bool:
        """删除成品减水剂"""
        data = self.load_data()
        products = data.get(DataCategory.PRODUCTS.value, [])
        
        new_products = [p for p in products if p.get("id") != product_id]
        
        if len(new_products) < len(products):
            data[DataCategory.PRODUCTS.value] = new_products
            return self.save_data(data)
        return False

    def update_product(self, product_id: int, updated_fields: Dict[str, Any]) -> bool:
        """更新成品减水剂"""
        # 如果更新了名称，检查唯一性
        if "product_name" in updated_fields:
             if self.check_product_name_exists(updated_fields["product_name"], exclude_id=product_id):
                 raise ValueError(f"成品名称 '{updated_fields['product_name']}' 已存在")

        data = self.load_data()
        products = data.get(DataCategory.PRODUCTS.value, [])
        updated = False
        for i, product in enumerate(products):
            if product.get("id") == product_id:
                new_product = dict(product)
                new_product.update(updated_fields)
                new_product["id"] = product_id
                if "created_at" not in new_product and product.get("created_at"):
                    new_product["created_at"] = product.get("created_at")
                new_product["last_modified"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                products[i] = new_product
                updated = True
                break
        if updated:
            data[DataCategory.PRODUCTS.value] = products
            return self.save_data(data)
        return False

    # -------------------- BOM 管理 (M2) --------------------
    # Ensure this section is loaded
    def get_all_boms(self):
        """获取所有BOM主数据"""
        data = self.load_data()
        return data.get(DataCategory.BOMS.value, [])

    def add_bom(self, bom_data: Union[Dict[str, Any], BOM]) -> Optional[int]:
        """添加BOM主数据"""
        data = self.load_data()
        boms = data.get(DataCategory.BOMS.value, [])
        
        new_id = self._get_next_id(boms)
        
        if isinstance(bom_data, BOM):
            bom_data.id = new_id
            if not bom_data.created_at:
                bom_data.created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            if not bom_data.last_modified:
                bom_data.last_modified = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            final_bom = bom_data.model_dump(mode='json')
        else:
            bom_data["id"] = new_id
            if "created_at" not in bom_data:
                bom_data["created_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            if "last_modified" not in bom_data:
                bom_data["last_modified"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            try:
                BOM(**bom_data)
            except Exception as e:
                logger.warning(f"BOM validation warning: {e}")
            final_bom = bom_data
        
        boms.append(final_bom)
        data[DataCategory.BOMS.value] = boms
        if self.save_data(data):
            return new_id
        return None

    def update_bom(self, bom_id: int, updated_fields: Dict[str, Any]) -> bool:
        """更新BOM主数据"""
        data = self.load_data()
        boms = data.get(DataCategory.BOMS.value, [])
        updated = False
        for i, bom in enumerate(boms):
            if bom.get("id") == bom_id:
                boms[i].update(updated_fields)
                boms[i]["last_modified"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                updated = True
                break
        if updated:
            data[DataCategory.BOMS.value] = boms
            return self.save_data(data)
        return False
        
    def delete_bom(self, bom_id: int) -> bool:
        """删除BOM（软删除或硬删除，这里做硬删除但要检查版本）"""
        # 实际业务中建议检查是否有生产单引用
        data = self.load_data()
        boms = data.get(DataCategory.BOMS.value, [])
        versions = data.get(DataCategory.BOM_VERSIONS.value, [])
        
        # 级联删除版本
        new_versions = [v for v in versions if v.get("bom_id") != bom_id]
        new_boms = [b for b in boms if b.get("id") != bom_id]
        
        if len(new_boms) < len(boms):
            data[DataCategory.BOMS.value] = new_boms
            data[DataCategory.BOM_VERSIONS.value] = new_versions
            return self.save_data(data)
        return False

    def get_bom_versions(self, bom_id: int) -> List[Dict[str, Any]]:
        """获取指定BOM的所有版本"""
        data = self.load_data()
        versions = data.get(DataCategory.BOM_VERSIONS.value, [])
        return [v for v in versions if v.get("bom_id") == bom_id]

    def get_all_bom_versions(self) -> List[Dict[str, Any]]:
        data = self.load_data()
        return data.get(DataCategory.BOM_VERSIONS.value, [])

    def get_effective_bom_version(self, bom_id: int, as_of_date: Optional[date] = None) -> Optional[Dict[str, Any]]:
        """根据生效日期获取指定BOM在某一时点的有效版本
        
        策略:
        - 只考虑有明细行的版本
        - 只考虑 effective_from 不晚于 as_of_date 的版本
        - 按 effective_from 倒序，若相同则按 id 倒序，选最新一条
        """
        versions = self.get_bom_versions(bom_id)
        if not versions:
            return None

        if as_of_date is None:
            as_of_date = datetime.now().date()

        def _parse_date(v):
            d = v.get("effective_from")
            if not d:
                return None
            try:
                return datetime.strptime(str(d), "%Y-%m-%d").date()
            except Exception:
                return None

        candidates = []
        for v in versions:
            status = v.get("status")
            if status in ["pending", "rejected"]:
                continue
            if not v.get("lines"):
                continue
            eff = _parse_date(v)
            if eff is None or eff > as_of_date:
                continue
            candidates.append((eff, int(v.get("id", 0)), v))

        if not candidates:
            with_lines = []
            for v in versions:
                status = v.get("status")
                if status in ["pending", "rejected"]:
                    continue
                if v.get("lines"):
                    with_lines.append(v)
            if not with_lines:
                return None
            return sorted(with_lines, key=lambda x: int(x.get("id", 0)), reverse=True)[0]

        candidates.sort(key=lambda x: (x[0], x[1]), reverse=True)
        return candidates[0][2]

    def add_bom_version(self, version_data: Union[Dict[str, Any], BOMVersion]) -> Optional[int]:
        """添加BOM版本"""
        data = self.load_data()
        versions = data.get(DataCategory.BOM_VERSIONS.value, [])
        
        user = None
        try:
            user = st.session_state.get("current_user")
        except Exception:
            user = None

        new_id = self._get_next_id(versions)
        
        if isinstance(version_data, BOMVersion):
            version_data.id = new_id
            if not version_data.created_at:
                version_data.created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            if user:
                if not version_data.created_by:
                    version_data.created_by = user.get("username")
            
            # Admin auto-approve logic
            if version_data.status == BOMStatus.PENDING:
                 if user and user.get("role") == UserRole.ADMIN.value:
                      version_data.status = BOMStatus.ACTIVE

            final_ver = version_data.model_dump(mode='json')
            if user:
                final_ver["created_role"] = user.get("role")
        else:
            if "status" not in version_data:
                if user and user.get("role") == UserRole.ADMIN.value:
                    version_data["status"] = BOMStatus.ACTIVE.value
                else:
                    version_data["status"] = BOMStatus.PENDING.value
            if user and "created_by" not in version_data:
                version_data["created_by"] = user.get("username")
                version_data["created_role"] = user.get("role")

            version_data["id"] = new_id
            if "created_at" not in version_data:
                version_data["created_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            try:
                BOMVersion(**version_data)
            except Exception as e:
                logger.warning(f"BOMVersion validation warning: {e}")
            final_ver = version_data
        
        versions.append(final_ver)
        data[DataCategory.BOM_VERSIONS.value] = versions
        if self.save_data(data):
            return new_id
        return None

    def update_bom_version(self, version_id: int, updated_fields: Dict[str, Any]) -> bool:
        """更新BOM版本"""
        data = self.load_data()
        versions = data.get(DataCategory.BOM_VERSIONS.value, [])
        updated = False
        for i, v in enumerate(versions):
            if v.get("id") == version_id:
                versions[i].update(updated_fields)
                updated = True
                break
        if updated:
            data[DataCategory.BOM_VERSIONS.value] = versions
            return self.save_data(data)
        return False

    def delete_bom_version(self, version_id: int) -> Tuple[bool, str]:
        data = self.load_data()
        versions = data.get(DataCategory.BOM_VERSIONS.value, [])
        orders = data.get(DataCategory.PRODUCTION_ORDERS.value, [])
        target = None
        for v in versions:
            if v.get("id") == version_id:
                target = v
                break
        if not target:
            return False, "BOM版本不存在"
        for o in orders:
            if o.get("bom_version_id") == version_id:
                return False, "存在引用该版本的生产单，无法删除"
        new_versions = [v for v in versions if v.get("id") != version_id]
        data[DataCategory.BOM_VERSIONS.value] = new_versions
        if self.save_data(data):
            return True, "删除成功"
        return False, "保存失败"

    def explode_bom(self, bom_version_id: Union[int, str], target_qty: float = 1000.0) -> List[Dict[str, Any]]:
        """
        BOM 展开计算
        Args:
            bom_version_id: 版本ID
            target_qty: 目标产量
        Returns:
            list: [{item_id, item_name, item_type, required_qty, uom, ...}]
        """
        data = self.load_data()
        versions = data.get(DataCategory.BOM_VERSIONS.value, [])
        
        # 兼容 ID 类型比较 (int vs str)
        target_version = None
        for v in versions:
            if str(v.get("id")) == str(bom_version_id):
                target_version = v
                break
        
        if not target_version:
            logger.warning(f"Explode BOM failed: Version ID {bom_version_id} not found.")
            return []
            
        base_qty = float(target_version.get("yield_base", 1000.0) or 1000.0)
        if base_qty <= 0: base_qty = 1000.0
        
        ratio = target_qty / base_qty
        
        result = []
        lines = target_version.get("lines", [])
        
        if not lines:
            logger.warning(f"Explode BOM: Version {bom_version_id} has no lines.")
            
        for line in lines:
            qty = float(line.get("qty", 0.0) or 0.0) * ratio
            item_type = line.get("item_type")
            item_id = line.get("item_id")
            
            # 基础信息
            item_info = {
                "item_id": item_id,
                "item_type": item_type,
                "required_qty": qty,
                "uom": line.get("uom", UnitType.KG.value),
                "phase": line.get("phase", ""),
                "item_name": line.get("item_name", "Unknown")
            }
            
            result.append(item_info)
            
        return result
    
    # -------------------- 生产与库存扩展 (M3) --------------------
    def get_stock_balance(self, material_id: Optional[int] = None) -> Union[float, Dict[int, float]]:
        """获取库存余额（通过汇总台账）"""
        data = self.load_data()
        records = data.get(DataCategory.INVENTORY_RECORDS.value, [])
        
        # 如果指定了 material_id，只计算该物料
        if material_id:
            balance = 0.0
            for r in records:
                if r.get("material_id") == material_id:
                    qty = float(r.get("quantity", 0.0))
                    rtype = r.get("type", "")
                    if rtype in [StockMovementType.IN.value, StockMovementType.PRODUCE_IN.value, StockMovementType.ADJUST_IN.value, StockMovementType.RETURN_IN.value]:
                        balance += qty
                    elif rtype in [StockMovementType.OUT.value, StockMovementType.CONSUME_OUT.value, StockMovementType.ADJUST_OUT.value]:
                        balance -= qty
            return balance
        
        # 否则返回所有物料的余额字典
        balances = {}
        for r in records:
            mid = r.get("material_id")
            if mid not in balances: balances[mid] = 0.0
            
            qty = float(r.get("quantity", 0.0))
            rtype = r.get("type", "")
            if rtype in [StockMovementType.IN.value, StockMovementType.PRODUCE_IN.value, StockMovementType.ADJUST_IN.value, StockMovementType.RETURN_IN.value]:
                balances[mid] += qty
            elif rtype in [StockMovementType.OUT.value, StockMovementType.CONSUME_OUT.value, StockMovementType.ADJUST_OUT.value]:
                balances[mid] -= qty
        return balances

    def add_production_order(self, order_data: Union[Dict[str, Any], ProductionOrder]) -> Optional[int]:
        """添加生产单"""
        data = self.load_data()
        orders = data.get(DataCategory.PRODUCTION_ORDERS.value, [])
        
        new_id = self._get_next_id(orders)
        
        if isinstance(order_data, ProductionOrder):
            order_data.id = new_id
            if not order_data.created_at:
                order_data.created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            final_order = order_data.model_dump(mode='json')
        else:
            order_data["id"] = new_id
            if "created_at" not in order_data:
                order_data["created_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            try:
                ProductionOrder(**order_data)
            except Exception as e:
                logger.warning(f"ProductionOrder validation warning: {e}")
            final_order = order_data
        
        orders.append(final_order)
        data[DataCategory.PRODUCTION_ORDERS.value] = orders
        if self.save_data(data):
            return new_id
        return None

    def update_production_order(self, order_id: int, updated_fields: Dict[str, Any]) -> bool:
        """更新生产单"""
        data = self.load_data()
        orders = data.get(DataCategory.PRODUCTION_ORDERS.value, [])
        updated = False
        for i, order in enumerate(orders):
            if order.get("id") == order_id:
                orders[i].update(updated_fields)
                orders[i]["last_modified"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                updated = True
                break
        if updated:
            data[DataCategory.PRODUCTION_ORDERS.value] = orders
            return self.save_data(data)
        return False
        
    def get_all_production_orders(self) -> List[Dict[str, Any]]:
        """获取所有生产单"""
        data = self.load_data()
        return data.get(DataCategory.PRODUCTION_ORDERS.value, [])

    def get_production_orders(self) -> List[Dict[str, Any]]:
        """获取所有生产单（别名方法，兼容调用）"""
        return self.get_all_production_orders()

    def delete_production_order(self, order_id: int) -> Tuple[bool, str]:
        """删除生产单（及其关联的领料单）"""
        data = self.load_data()
        orders = data.get(DataCategory.PRODUCTION_ORDERS.value, [])
        issues = data.get(DataCategory.MATERIAL_ISSUES.value, [])
        
        # 检查是否可以删除
        order = next((o for o in orders if o.get("id") == order_id), None)
        if not order: return False, "生产单不存在"
        
        # 如果已经有领料单过账，则不允许删除
        related_issues = [i for i in issues if i.get("production_order_id") == order_id]
        for issue in related_issues:
            if issue.get("status") == IssueStatus.POSTED.value:
                return False, "无法删除：存在已过账的领料单"
        
        # 删除生产单
        new_orders = [o for o in orders if o.get("id") != order_id]
        # 删除关联的草稿领料单
        new_issues = [i for i in issues if i.get("production_order_id") != order_id]
        
        if len(new_orders) < len(orders):
            data[DataCategory.PRODUCTION_ORDERS.value] = new_orders
            data[DataCategory.MATERIAL_ISSUES.value] = new_issues
            if self.save_data(data):
                return True, "删除成功"
            return False, "保存失败"
        return False, "删除未生效"

    def finish_production_order(self, order_id: int, operator: str = "User") -> Tuple[bool, str]:
        """
        完工入库：
        1. 更新生产单状态为 finished
        2. 根据 BOM 对应的成品，增加成品库存
        3. 记录成品入库流水
        """
        data = self.load_data()
        orders = data.get(DataCategory.PRODUCTION_ORDERS.value, [])
        inventory = data.get(DataCategory.PRODUCT_INVENTORY.value, [])
        records = data.get(DataCategory.PRODUCT_INVENTORY_RECORDS.value, [])
        boms = data.get(DataCategory.BOMS.value, [])
        
        target_order = None
        target_idx = -1
        for i, o in enumerate(orders):
            if o.get("id") == order_id:
                target_order = o
                target_idx = i
                break
        
        if not target_order: return False, "生产单不存在"
        if target_order.get("status") == ProductionOrderStatus.FINISHED.value: return False, "生产单已完工"
        
        # 获取计划产量 (这里简化为实际产量=计划产量，实际应用可能需要输入实际产量)
        plan_qty = float(target_order.get("plan_qty", 0.0))
        if plan_qty <= 0: return False, "计划产量无效"
        
        # 查找对应的 BOM 信息以确定产品名称和类型
        bom_id = target_order.get("bom_id")
        target_bom = next((b for b in boms if b.get("id") == bom_id), None)
        
        if not target_bom: return False, "关联 BOM 不存在"
        
        # 构造产品名称: {bom_code}-{bom_name}
        bom_code = target_bom.get("bom_code", "").strip()
        bom_name = target_bom.get("bom_name", "").strip()
        
        if bom_code:
            product_name = f"{bom_code}-{bom_name}"
        else:
            product_name = bom_name

        # 映射 BOM 类型到产品类型
        # BOM: 母液, 成品, 速凝剂, 防冻剂
        # Product: 母液, 成品, 速凝剂, 防冻剂, 其他
        product_type = target_bom.get("bom_type", "其他")
        
        # 在 product_inventory 中查找对应产品
        target_prod_idx = -1
        
        # 智能匹配：尝试匹配带前缀和不带前缀的名称，以解决 "WJSNJ-无碱速凝剂" 与 "无碱速凝剂" 分离的问题
        candidate_names = [product_name]
        if "-" in product_name:
            # 尝试添加不带前缀的名称 (如 "无碱速凝剂")
            candidate_names.append(product_name.split("-", 1)[1])
        
        # 同样尝试添加带标准前缀的名称 (如果当前是不带前缀的)
        if product_name == "无碱速凝剂":
            candidate_names.append("WJSNJ-无碱速凝剂")
        elif product_name == "有碱速凝剂":
            candidate_names.append("YJSNJ-有碱速凝剂")
            
        for i, p in enumerate(inventory):
            # 优先匹配名称，类型作为辅助验证 (或忽略类型差异，因为名称通常是唯一的)
            if p.get("name") in candidate_names:
                target_prod_idx = i
                break
        
        # 如果产品不存在，自动创建
        if target_prod_idx == -1:
            new_prod_id = self._get_next_id(inventory)
            new_prod = {
                "id": new_prod_id,
                "name": product_name,
                "type": product_type,
                "stock_quantity": 0.0,
                "unit": "kg", # AI_RULES: 强制 kg
                "last_update": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            }
            inventory.append(new_prod)
            target_prod_idx = len(inventory) - 1
        
        # 更新库存
        # AI_RULES: 生产单 plan_qty 已经是 kg，库存也是 kg，无需转换
        final_qty = plan_qty
        
        current_stock = float(inventory[target_prod_idx].get("stock_quantity", 0.0))
        new_stock = current_stock + final_qty
        inventory[target_prod_idx]["stock_quantity"] = new_stock
        inventory[target_prod_idx]["last_update"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # 记录流水
        new_rec_id = self._get_next_id(records)
        records.append({
            "id": new_rec_id,
            "date": datetime.now().strftime("%Y-%m-%d"),
            "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "product_name": inventory[target_prod_idx]["name"], # 使用库存中实际的产品名称，确保关联正确
            "product_type": product_type,
            "type": StockMovementType.PRODUCE_IN.value, # 明确标记为生产入库
            "quantity": final_qty, # kg
            "reason": f"生产完工: {target_order.get('order_code')}",
            "operator": operator,
            "snapshot_stock": new_stock,
            "batch_number": target_order.get('order_code') # 使用生产单号作为批次号
        })
        
        # 更新订单状态
        orders[target_idx]["status"] = ProductionOrderStatus.FINISHED.value
        orders[target_idx]["finished_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        data[DataCategory.PRODUCTION_ORDERS.value] = orders
        data[DataCategory.PRODUCT_INVENTORY.value] = inventory
        data[DataCategory.PRODUCT_INVENTORY_RECORDS.value] = records
        
        if self.save_data(data):
            return True, f"完工入库成功，库存增加 {final_qty:.3f} kg"
        return False, "保存失败"

    def create_issue_from_order(self, order_id: int) -> Optional[int]:
        """
        根据生产单创建领料单（建议）
        Returns:
            int: 新领料单ID
        """
        data = self.load_data()
        orders = data.get(DataCategory.PRODUCTION_ORDERS.value, [])
        order = next((o for o in orders if o.get("id") == order_id), None)
        
        if not order: return None
        
        # 1. 找到对应的 BOM 版本
        bom_version_id = order.get("bom_version_id")
        bom_id = order.get("bom_id")
        plan_qty = float(order.get("plan_qty", 0.0))
        
        # 2. 展开 BOM（若无明细，回退到最近有效版本）
        lines = self.explode_bom(bom_version_id, plan_qty)
        if not lines:
            versions = data.get(DataCategory.BOM_VERSIONS.value, [])
            same_bom_versions = []
            for v in versions:
                if str(v.get("bom_id")) != str(bom_id):
                    continue
                status = v.get("status")
                if status in ["pending", "rejected"]:
                    continue
                same_bom_versions.append(v)
            fallback_ver = None
            for v in reversed(same_bom_versions):
                if v.get("lines"):
                    fallback_ver = v
                    break
            if fallback_ver:
                # 更新订单所选版本为回退版本
                for i, o in enumerate(orders):
                    if o.get("id") == order_id:
                        orders[i]["bom_version_id"] = fallback_ver["id"]
                        orders[i]["last_modified"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        break
                data[DataCategory.PRODUCTION_ORDERS.value] = orders
                # 重新展开
                lines = self.explode_bom(fallback_ver["id"], plan_qty)
            else:
                return None
        
        # 3. 生成领料单
        issues = data.get(DataCategory.MATERIAL_ISSUES.value, [])
        new_id = self._get_next_id(issues)
        
        # 确保行项目包含 item_type
        for line in lines:
            if "item_type" not in line:
                line["item_type"] = MaterialType.RAW_MATERIAL.value # 默认值，兼容旧数据
        
        issue_data = {
            "id": new_id,
            "issue_code": f"ISS-{datetime.now().strftime('%Y%m%d')}-{new_id:04d}",
            "production_order_id": order_id,
            "status": IssueStatus.DRAFT.value,
            "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "lines": lines  # 直接使用展开的行作为建议领料行
        }
        
        issues.append(issue_data)
        data[DataCategory.MATERIAL_ISSUES.value] = issues
        
        if self.save_data(data):
            return new_id
        return None

    def get_material_issues(self, order_id: Optional[int] = None) -> List[Dict[str, Any]]:
        """获取领料单"""
        data = self.load_data()
        issues = data.get(DataCategory.MATERIAL_ISSUES.value, [])
        if order_id:
            return [i for i in issues if i.get("production_order_id") == order_id]
        return issues
        
    def update_material_issue(self, issue_id: int, updated_fields: Dict[str, Any]) -> bool:
        """更新领料单（通常是更新实发数量）"""
        data = self.load_data()
        issues = data.get(DataCategory.MATERIAL_ISSUES.value, [])
        updated = False
        for i, issue in enumerate(issues):
            if issue.get("id") == issue_id:
                issues[i].update(updated_fields)
                updated = True
                break
        if updated:
            data[DataCategory.MATERIAL_ISSUES.value] = issues
            return self.save_data(data)
        return False

    def post_issue(self, issue_id: int, operator: str = "System") -> Tuple[bool, str]:
        """
        领料过账：
        1. 锁定领料单 status=posted
        2. 生成 inventory_records (type=consume_out)
        3. 更新库存 (支持原材料和产品)
        """
        data = self.load_data()
        issues = data.get(DataCategory.MATERIAL_ISSUES.value, [])
        orders = data.get(DataCategory.PRODUCTION_ORDERS.value, [])
        target_issue = None
        for issue in issues:
            if issue.get("id") == issue_id:
                target_issue = issue
                break
        
        if not target_issue: return False, "领料单不存在"
        if target_issue.get("status") == IssueStatus.POSTED.value: return False, "领料单已过账"
        if not target_issue.get("lines"):
            return False, "领料单没有明细，无法过账"
        
        # 准备批量写入台账
        records = data.get(DataCategory.INVENTORY_RECORDS.value, [])
        product_records = data.get(DataCategory.PRODUCT_INVENTORY_RECORDS.value, [])
        materials = data.get(DataCategory.RAW_MATERIALS.value, [])
        products = data.get(DataCategory.PRODUCT_INVENTORY.value, [])
        
        # 遍历行项目
        for line in target_issue.get("lines", []):
            qty = float(line.get("required_qty", 0.0))
            if qty <= 0: continue
            
            mid = line.get("item_id")
            line_uom = line.get("uom", UnitType.KG.value)
            item_type = line.get("item_type", MaterialType.RAW_MATERIAL.value)
            
            if item_type == MaterialType.PRODUCT.value:
                current_stock = 0.0
                prod_idx = -1
                expected_name = str(line.get("item_name", "") or "").strip()
                desired_name = None
                desired_type = None

                if expected_name:
                    if expected_name.startswith("YJSNJ-") or expected_name in ["有碱速凝剂", "碱速凝剂"]:
                        desired_name = "YJSNJ-有碱速凝剂"
                        desired_type = "有碱速凝剂"
                    elif expected_name.startswith("WJSNJ-") or expected_name in ["无碱速凝剂"]:
                        desired_name = "WJSNJ-无碱速凝剂"
                        desired_type = "无碱速凝剂"
                    else:
                        desired_name = expected_name
                
                for idx, p in enumerate(products):
                    if p.get("id") == mid:
                        current_stock = float(p.get("stock_quantity", 0.0))
                        prod_idx = idx
                        break
                
                corrected = False

                if prod_idx >= 0:
                    actual_name = str(products[prod_idx].get("name", "") or "").strip()
                    actual_type = str(products[prod_idx].get("type", "其他") or "").strip()
                    if not desired_name:
                        desired_name = actual_name
                    if not desired_type:
                        desired_type = actual_type
                    
                    # 检查是否需要切换产品 (例如 ID 对应的产品名称不符合预期)
                    # 增加别名匹配逻辑：如果名称虽不同但属于同一产品的不同别名 (如 "无碱速凝剂" 和 "WJSNJ-无碱速凝剂")，则视为匹配
                    is_alias_match = False
                    if ("无碱速凝剂" in actual_name and "无碱速凝剂" in desired_name) or \
                       ("有碱速凝剂" in actual_name and "有碱速凝剂" in desired_name):
                        is_alias_match = True
                        
                    if (actual_name != desired_name or actual_type != desired_type) and not is_alias_match:
                        new_idx = -1
                        for j, pj in enumerate(products):
                            if str(pj.get("name", "") or "").strip() == desired_name:
                                new_idx = j
                                break
                        if new_idx == -1 and desired_name:
                            new_id = self._get_next_id(products)
                            products.append({
                                "id": new_id,
                                "name": desired_name,
                                "type": desired_type or actual_type,
                                "stock_quantity": 0.0,
                                "unit": products[prod_idx].get("unit", "吨"),
                                "last_update": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                            })
                            new_idx = len(products) - 1
                        if new_idx != -1:
                            prod_idx = new_idx
                            current_stock = float(products[prod_idx].get("stock_quantity", 0.0))
                            corrected = True
                else:
                    if desired_name:
                        new_idx = -1
                        for j, pj in enumerate(products):
                            if str(pj.get("name", "") or "").strip() == desired_name:
                                new_idx = j
                                break
                        if new_idx == -1:
                            new_id = self._get_next_id(products)
                            products.append({
                                "id": new_id,
                                "name": desired_name,
                                "type": desired_type or "其他",
                                "stock_quantity": 0.0,
                                "unit": "吨",
                                "last_update": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                            })
                            new_idx = len(products) - 1
                        prod_idx = new_idx
                        current_stock = float(products[prod_idx].get("stock_quantity", 0.0))
                        corrected = True

                if prod_idx == -1:
                    continue

                stock_unit = products[prod_idx].get("unit", "吨")
                final_qty, success = convert_quantity(qty, line_uom, stock_unit)
                
                if not success and normalize_unit(line_uom) != normalize_unit(stock_unit):
                     logger.warning(f"Unit conversion failed in post_issue (product): {qty} {line_uom} -> {stock_unit}")

                new_stock = current_stock - final_qty
                products[prod_idx]["stock_quantity"] = new_stock
                products[prod_idx]["last_update"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                
                reason_note = f"生产领料: {target_issue.get('issue_code')}"
                if expected_name and corrected:
                    reason_note += f" (纠正: {expected_name})"
                if normalize_unit(line_uom) != normalize_unit(stock_unit):
                    reason_note += f" (原: {qty}{line_uom})"
                
                new_rec_id = self._get_next_id(product_records)
                rel_order_id = target_issue.get("production_order_id")
                rel_bom_id = None
                rel_bom_ver = None
                if rel_order_id:
                    ord = next((o for o in orders if o.get("id") == rel_order_id), None)
                    if ord:
                        rel_bom_id = ord.get("bom_id")
                        rel_bom_ver = ord.get("bom_version_id")
                product_records.append({
                    "id": new_rec_id,
                    "date": datetime.now().strftime("%Y-%m-%d"),
                    "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "product_name": products[prod_idx]["name"],
                    "product_type": products[prod_idx].get("type", "其他"),
                    "type": StockMovementType.CONSUME_OUT.value, # 标记为生产消耗
                    "quantity": final_qty,
                    "reason": reason_note,
                    "operator": operator,
                    "snapshot_stock": new_stock,
                    "related_doc_type": "ISSUE",
                    "related_doc_id": issue_id,
                    "related_order_id": rel_order_id,
                    "related_bom_id": rel_bom_id,
                    "related_bom_version_id": rel_bom_ver
                })
            else:
                # 处理原材料库存扣减 (原有逻辑)
                current_stock = 0.0
                mat_idx = -1
                is_water = False
                
                for idx, m in enumerate(materials):
                    if m.get("id") == mid:
                        current_stock = float(m.get("stock_quantity", 0.0))
                        mat_name = m.get("name", "").strip()
                        water_names = ["水", "自来水", "纯水", "去离子水", "工业用水", "生产用水"]
                        is_water = mat_name in water_names
                        mat_idx = idx
                        break
                
                if mat_idx >= 0:
                    # 单位转换
                    stock_unit = materials[mat_idx].get("unit", UnitType.KG.value)
                    final_qty, success = convert_quantity(qty, line_uom, stock_unit)
                    
                    if not success and normalize_unit(line_uom) != normalize_unit(stock_unit):
                         logger.warning(f"Unit conversion failed in post_issue: {qty} {line_uom} -> {stock_unit}")
    
                    new_stock = current_stock
                    if not is_water:
                        new_stock = current_stock - final_qty
                        materials[mat_idx]["stock_quantity"] = new_stock
                        materials[mat_idx]["last_stock_update"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    
                    # 写入台账 (记录转换后的数量，并备注原始数量)
                    reason_note = f"生产领料: {target_issue.get('issue_code')}"
                    if normalize_unit(line_uom) != normalize_unit(stock_unit):
                        reason_note += f" (原: {qty}{line_uom})"
                    
                    new_rec_id = self._get_next_id(records)
                    records.append({
                        "id": new_rec_id,
                        "date": datetime.now().strftime("%Y-%m-%d"),
                        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "material_id": mid,
                        "type": StockMovementType.CONSUME_OUT.value,
                        "quantity": final_qty,
                        "reason": reason_note,
                        "operator": operator,
                        "related_doc_type": "ISSUE",
                        "related_doc_id": issue_id,
                        "snapshot_stock": new_stock
                    })
        
        # 更新领料单状态
        target_issue["status"] = IssueStatus.POSTED.value
        target_issue["posted_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        data["inventory_records"] = records
        data["product_inventory_records"] = product_records
        data["raw_materials"] = materials
        data["product_inventory"] = products
        # material_issues 已经在 target_issue 引用中修改了
        
        if self.save_data(data):
            return True, "过账成功"
        return False, "保存失败"

    def cancel_issue_posting(self, issue_id: int, operator: str = "System") -> Tuple[bool, str]:
        """
        撤销领料过账：
        1. 检查状态是否为 posted
        2. 生成冲销台账 (type=return_in, reason=撤销领料)
        3. 恢复库存
        4. 状态回滚为 draft
        """
        data = self.load_data()
        issues = data.get(DataCategory.MATERIAL_ISSUES.value, [])
        orders = data.get(DataCategory.PRODUCTION_ORDERS.value, [])
        products = data.get(DataCategory.PRODUCT_INVENTORY.value, [])
        product_records = data.get(DataCategory.PRODUCT_INVENTORY_RECORDS.value, [])
        target_issue = None
        for issue in issues:
            if issue.get("id") == issue_id:
                target_issue = issue
                break
        
        if not target_issue: return False, "领料单不存在"
        if target_issue.get("status") != IssueStatus.POSTED.value: return False, "只有已过账的领料单可以撤销"
        
        records = data.get(DataCategory.INVENTORY_RECORDS.value, [])
        materials = data.get(DataCategory.RAW_MATERIALS.value, [])
        
        # 遍历行项目进行回滚
        for line in target_issue.get("lines", []):
            qty = float(line.get("required_qty", 0.0))
            if qty <= 0: continue
            
            mid = line.get("item_id")
            line_uom = line.get("uom", UnitType.KG.value)
            item_type = line.get("item_type", MaterialType.RAW_MATERIAL.value)
            if item_type == MaterialType.PRODUCT.value:
                prod_idx = -1
                current_stock = 0.0
                expected_name = str(line.get("item_name", "") or "").strip()
                desired_name = None
                desired_type = None
                if expected_name:
                    if expected_name.startswith("YJSNJ-") or expected_name in ["有碱速凝剂", "碱速凝剂"]:
                        desired_name = "YJSNJ-有碱速凝剂"
                        desired_type = "有碱速凝剂"
                    elif expected_name.startswith("WJSNJ-") or expected_name in ["无碱速凝剂"]:
                        desired_name = "WJSNJ-无碱速凝剂"
                        desired_type = "无碱速凝剂"
                    else:
                        desired_name = expected_name
                for idx, p in enumerate(products):
                    if p.get("id") == mid:
                        prod_idx = idx
                        current_stock = float(p.get("stock_quantity", 0.0))
                        break
                if prod_idx == -1 and desired_name:
                    for idx, p in enumerate(products):
                        if str(p.get("name", "") or "").strip() == desired_name:
                            prod_idx = idx
                            current_stock = float(p.get("stock_quantity", 0.0))
                            break
                if prod_idx >= 0:
                    stock_unit = products[prod_idx].get("unit", UnitType.TON.value)
                    final_qty, success = convert_quantity(qty, line_uom, stock_unit)
                    if not success and normalize_unit(line_uom) != normalize_unit(stock_unit):
                        logger.warning(f"Unit conversion failed in cancel_issue (product): {qty} {line_uom} -> {stock_unit}")
                    new_stock = current_stock + final_qty
                    products[prod_idx]["stock_quantity"] = new_stock
                    products[prod_idx]["last_update"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    
                    reason_note = f"撤销领料: {target_issue.get('issue_code')}"
                    if normalize_unit(line_uom) != normalize_unit(stock_unit):
                        reason_note += f" (原: {qty}{line_uom})"
                    new_rec_id = self._get_next_id(product_records)
                    rel_order_id = target_issue.get("production_order_id")
                    rel_bom_id = None
                    rel_bom_ver = None
                    if rel_order_id:
                        ord = next((o for o in orders if o.get("id") == rel_order_id), None)
                        if ord:
                            rel_bom_id = ord.get("bom_id")
                            rel_bom_ver = ord.get("bom_version_id")
                    product_records.append({
                        "id": new_rec_id,
                        "date": datetime.now().strftime("%Y-%m-%d"),
                        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "product_name": products[prod_idx].get("name"),
                        "product_type": products[prod_idx].get("type", "其他"),
                        "type": StockMovementType.RETURN_IN.value,
                        "quantity": final_qty,
                        "reason": reason_note,
                        "operator": operator,
                        "snapshot_stock": new_stock,
                        "related_doc_type": "ISSUE_CANCEL",
                        "related_doc_id": issue_id,
                        "related_order_id": rel_order_id,
                        "related_bom_id": rel_bom_id,
                        "related_bom_version_id": rel_bom_ver
                    })
            else:
                # 更新原材料库存（加回）
                current_stock = 0.0
                mat_idx = -1
                is_water = False
                
                for idx, m in enumerate(materials):
                    if m.get("id") == mid:
                        current_stock = float(m.get("stock_quantity", 0.0))
                        mat_name = m.get("name", "").strip()
                        water_names = ["水", "自来水", "纯水", "去离子水", "工业用水", "生产用水"]
                        is_water = mat_name in water_names
                        mat_idx = idx
                        break
                
                if mat_idx >= 0:
                    stock_unit = materials[mat_idx].get("unit", UnitType.KG.value)
                    final_qty, success = convert_quantity(qty, line_uom, stock_unit)

                    if not success and normalize_unit(line_uom) != normalize_unit(stock_unit):
                         logger.warning(f"Unit conversion failed in cancel_issue: {qty} {line_uom} -> {stock_unit}")

                    new_stock = current_stock
                    if not is_water:
                        new_stock = current_stock + final_qty
                        materials[mat_idx]["stock_quantity"] = new_stock
                        materials[mat_idx]["last_stock_update"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    
                    reason_note = f"撤销领料: {target_issue.get('issue_code')}"
                    if normalize_unit(line_uom) != normalize_unit(stock_unit):
                        reason_note += f" (原: {qty}{line_uom})"

                    new_rec_id = self._get_next_id(records)
                    records.append({
                        "id": new_rec_id,
                        "date": datetime.now().strftime("%Y-%m-%d"),
                        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "material_id": mid,
                        "type": StockMovementType.RETURN_IN.value,
                        "quantity": final_qty,
                        "reason": reason_note,
                        "operator": operator,
                        "related_doc_type": "ISSUE_CANCEL",
                        "related_doc_id": issue_id,
                        "snapshot_stock": new_stock
                    })
        
        # 状态回滚
        target_issue["status"] = IssueStatus.DRAFT.value
        target_issue["last_modified"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # 标记此前与该领料单关联的成品出库记录为已撤销
        for r in product_records:
            if (
                r.get("type") == StockMovementType.OUT.value
                and r.get("related_doc_type") == "ISSUE"
                and r.get("related_doc_id") == issue_id
            ):
                r["related_doc_type"] = "ISSUE_CANCEL"
                if r.get("reason"):
                    r["reason"] = f"{r['reason']} (已撤销)"
                else:
                    r["reason"] = "生产领料出库 (已撤销)"
        data[DataCategory.INVENTORY_RECORDS.value] = records
        data[DataCategory.RAW_MATERIALS.value] = materials
        data[DataCategory.PRODUCT_INVENTORY.value] = products
        data[DataCategory.PRODUCT_INVENTORY_RECORDS.value] = product_records
        
        if self.save_data(data):
            return True, "撤销成功，库存已恢复"
        return False, "保存失败"
    
    def repair_material_issues(self) -> bool:
        data = self.load_data()
        issues = data.get(DataCategory.MATERIAL_ISSUES.value, [])
        orders = data.get(DataCategory.PRODUCTION_ORDERS.value, [])
        versions = data.get(DataCategory.BOM_VERSIONS.value, [])
        updated = False
        for i, issue in enumerate(issues):
            if issue.get("status") != IssueStatus.DRAFT.value:
                continue
            oid = issue.get("production_order_id")
            order = next((o for o in orders if o.get("id") == oid), None)
            if not order:
                continue
            bom_ver_id = order.get("bom_version_id")
            bom_id = order.get("bom_id")
            plan_qty = float(order.get("plan_qty", 0.0))
            lines = self.explode_bom(bom_ver_id, plan_qty)
            if not lines:
                same_bom_versions = []
                for v in versions:
                    if str(v.get("bom_id")) != str(bom_id):
                        continue
                    status = v.get("status")
                    if status in [BOMStatus.PENDING.value, BOMStatus.REJECTED.value]:
                        continue
                    same_bom_versions.append(v)
                fallback_ver = None
                for v in reversed(same_bom_versions):
                    if v.get("lines"):
                        fallback_ver = v
                        break
                if fallback_ver:
                    for j, o in enumerate(orders):
                        if o.get("id") == oid:
                            orders[j]["bom_version_id"] = fallback_ver["id"]
                            orders[j]["last_modified"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                            break
                    data[DataCategory.PRODUCTION_ORDERS.value] = orders
                    lines = self.explode_bom(fallback_ver["id"], plan_qty)
            if lines:
                for line in lines:
                    if "item_type" not in line:
                        line["item_type"] = MaterialType.RAW_MATERIAL.value
                issues[i]["lines"] = lines
                issues[i]["last_modified"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                updated = True
        if updated:
            data[DataCategory.MATERIAL_ISSUES.value] = issues
            return self.save_data(data)
        return True
        
    # 净浆实验
    def get_all_paste_experiments(self) -> List[Dict[str, Any]]:
        """获取所有净浆实验"""
        data = self.load_data()
        return data.get(DataCategory.PASTE_EXPERIMENTS.value, [])
    
    def add_paste_experiment(self, experiment_data: Union[Dict[str, Any], PasteExperiment]) -> bool:
        """添加新净浆实验"""
        data = self.load_data()
        experiments = data.get(DataCategory.PASTE_EXPERIMENTS.value, [])
        
        # 生成新ID
        new_id = self._get_next_id(experiments)
        
        if isinstance(experiment_data, PasteExperiment):
            experiment_data.id = new_id
            if not experiment_data.created_at:
                experiment_data.created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            exp_dict = experiment_data.model_dump(mode='json')
        else:
            experiment_data["id"] = new_id
            if "created_at" not in experiment_data:
                experiment_data["created_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            try:
                PasteExperiment(**experiment_data)
            except Exception as e:
                logger.warning(f"PasteExperiment validation warning: {e}")
            exp_dict = experiment_data
        
        experiments.append(exp_dict)
        data[DataCategory.PASTE_EXPERIMENTS.value] = experiments
        return self.save_data(data)
    
    def delete_paste_experiment(self, experiment_id: int) -> bool:
        """删除净浆实验"""
        data = self.load_data()
        experiments = data.get(DataCategory.PASTE_EXPERIMENTS.value, [])
        
        new_experiments = [e for e in experiments if e.get("id") != experiment_id]
        
        if len(new_experiments) < len(experiments):
            data[DataCategory.PASTE_EXPERIMENTS.value] = new_experiments
            return self.save_data(data)
        return False

    def update_paste_experiment(self, experiment_id: int, updated_fields: Dict[str, Any]) -> bool:
        """更新净浆实验"""
        data = self.load_data()
        experiments = data.get(DataCategory.PASTE_EXPERIMENTS.value, [])
        updated = False
        for i, exp in enumerate(experiments):
            if exp.get("id") == experiment_id:
                experiments[i].update(updated_fields)
                experiments[i]["last_modified"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                updated = True
                break
        if updated:
            data[DataCategory.PASTE_EXPERIMENTS.value] = experiments
            return self.save_data(data)
        return False
    
    # 砂浆实验
    def get_all_mortar_experiments(self) -> List[Dict[str, Any]]:
        """获取所有砂浆实验"""
        data = self.load_data()
        return data.get(DataCategory.MORTAR_EXPERIMENTS.value, [])
    
    def add_mortar_experiment(self, experiment_data: Union[Dict[str, Any], MortarExperiment]) -> bool:
        """添加新砂浆实验"""
        data = self.load_data()
        experiments = data.get(DataCategory.MORTAR_EXPERIMENTS.value, [])
        
        # 生成新ID
        new_id = self._get_next_id(experiments)
        
        if isinstance(experiment_data, MortarExperiment):
            experiment_data.id = new_id
            if not experiment_data.created_at:
                experiment_data.created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            exp_dict = experiment_data.model_dump(mode='json')
        else:
            experiment_data["id"] = new_id
            if "created_at" not in experiment_data:
                experiment_data["created_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            try:
                MortarExperiment(**experiment_data)
            except Exception as e:
                logger.warning(f"MortarExperiment validation warning: {e}")
            exp_dict = experiment_data
            
        experiments.append(exp_dict)
        data[DataCategory.MORTAR_EXPERIMENTS.value] = experiments
        return self.save_data(data)
    
    def delete_mortar_experiment(self, experiment_id: int) -> bool:
        """删除砂浆实验"""
        data = self.load_data()
        experiments = data.get(DataCategory.MORTAR_EXPERIMENTS.value, [])
        
        new_experiments = [e for e in experiments if e.get("id") != experiment_id]
        
        if len(new_experiments) < len(experiments):
            data[DataCategory.MORTAR_EXPERIMENTS.value] = new_experiments
            return self.save_data(data)
        return False

    def update_mortar_experiment(self, experiment_id: int, updated_fields: Dict[str, Any]) -> bool:
        """更新砂浆实验"""
        data = self.load_data()
        experiments = data.get(DataCategory.MORTAR_EXPERIMENTS.value, [])
        updated = False
        for i, exp in enumerate(experiments):
            if exp.get("id") == experiment_id:
                experiments[i].update(updated_fields)
                experiments[i]["last_modified"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                updated = True
                break
        if updated:
            data[DataCategory.MORTAR_EXPERIMENTS.value] = experiments
            return self.save_data(data)
        return False

    # -----------------------------------------------------------
    # 产品库存管理 (Product Inventory)
    # -----------------------------------------------------------
    
    def get_product_inventory(self) -> List[Dict[str, Any]]:
        """获取所有产品库存"""
        data = self.load_data()
        return data.get("product_inventory", [])
        
    def add_product_inventory_record(self, record_data: Union[Dict[str, Any], ProductInventoryRecord], update_master_stock: bool = True) -> Tuple[bool, str]:
        """
        添加产品库存记录 (入库/出库)
        record_data: {
            "product_name": str,
            "product_type": str, (母液/速凝剂/防冻剂/减水剂/...)
            "quantity": float, (吨)
            "type": "in" | "out",
            "reason": str,
            "operator": str,
            "date": str
        }
        """
        data = self.load_data()
        inventory = data.get("product_inventory", [])
        records = data.get("product_inventory_records", [])
        
        # 统一转为字典处理
        if isinstance(record_data, ProductInventoryRecord):
            # 如果是模型，先确保ID和Created_at（如果有默认值逻辑的话）
            # 但这里ID是在后面生成的。
            rec_dict = record_data.model_dump(mode='json')
        else:
            rec_dict = record_data
            # 简单校验
            try:
                # 临时校验，不强制要求所有字段完全匹配，因为这里ID还没生成
                # ProductInventoryRecord(**rec_dict) 
                pass
            except Exception as e:
                logger.warning(f"ProductInventoryRecord validation warning: {e}")

        # 1. 更新库存快照
        product_name = rec_dict.get("product_name")
        product_type = rec_dict.get("product_type", ProductCategory.OTHER.value)
        qty = float(rec_dict.get("quantity", 0.0))
        op_type = rec_dict.get("type")
        
        target_item = next((p for p in inventory if p.get("product_name") == product_name), None)
        
        if not target_item:
            # 新增产品
            if op_type == StockMovementType.OUT.value:
                return False, "库存不足 (产品不存在)"
            
            target_item = {
                "id": self._get_next_id(inventory),
                "product_name": product_name,
                "type": product_type,
                "current_stock": qty,
                "unit": DEFAULT_UNIT,
                "last_update": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            }
            inventory.append(target_item)
        else:
            # 更新现有产品
            current_stock = float(target_item.get("current_stock", 0.0))
            
            if update_master_stock:
                if op_type in [StockMovementType.IN.value, StockMovementType.PRODUCE_IN.value, StockMovementType.RETURN_IN.value, StockMovementType.ADJUST_IN.value]:
                    current_stock += qty
                elif op_type in [StockMovementType.OUT.value, StockMovementType.SHIP_OUT.value, StockMovementType.CONSUME_OUT.value, StockMovementType.ADJUST_OUT.value]:
                    if current_stock < qty:
                        # 允许负库存? 暂时允许警告但继续，或者禁止
                        pass 
                    current_stock -= qty
                
                target_item["current_stock"] = current_stock
                target_item["last_update"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
        # 2. 添加流水记录
        new_rec_id = self._get_next_id(records)
        rec_dict["id"] = new_rec_id
        if "created_at" not in rec_dict:
            rec_dict["created_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        rec_dict["snapshot_stock"] = target_item["current_stock"]
        
        # 再次校验完整性（可选）
        if isinstance(record_data, ProductInventoryRecord):
             # 如果输入是对象，更新对象ID
             record_data.id = new_rec_id
             record_data.snapshot_stock = target_item["current_stock"]
             if not record_data.created_at:
                 record_data.created_at = rec_dict["created_at"]
        
        records.append(rec_dict)
        
        data["product_inventory"] = inventory
        data["product_inventory_records"] = records
        
        if self.save_data(data):
            return True, "库存更新成功"
        return False, "保存失败"

    def update_product_inventory_item(self, product_id: int, updates: Dict[str, Any]) -> Tuple[bool, str]:
        """更新产品信息（支持级联更新名称和类型）"""
        data = self.load_data()
        inventory = data.get("product_inventory", [])
        records = data.get("product_inventory_records", [])
        
        target_item = None
        target_idx = -1
        
        for i, item in enumerate(inventory):
            if item.get("id") == product_id:
                target_item = item
                target_idx = i
                break
        
        if target_item:
            # 检查是否需要级联更新
            old_name = target_item.get("product_name")
            old_type = target_item.get("type")
            new_name = updates.get("product_name", old_name)
            new_type = updates.get("type", old_type)
            
            need_cascade = (new_name != old_name) or (new_type != old_type)
            
            # 更新主数据
            inventory[target_idx].update(updates)
            inventory[target_idx]["last_update"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            if need_cascade:
                count = 0
                for record in records:
                    # 匹配名称和类型 (假设组合唯一)
                    # 注意：如果只改了 type，name没变，也得更新 type。
                    # 如果只改了 name，type没变，也得更新 name。
                    # 这里的匹配条件是 old_name 和 old_type。
                    if record.get("product_name") == old_name and record.get("product_type") == old_type:
                        record["product_name"] = new_name
                        record["product_type"] = new_type
                        count += 1
                if count > 0:
                    logger.info(f"Cascaded update for product {product_id}: {count} records updated.")

            data["product_inventory"] = inventory
            data["product_inventory_records"] = records # 确保记录也被保存
            if self.save_data(data):
                return True, "更新成功"
            return False, "保存失败"
        return False, "未找到指定产品"
            
    def get_product_inventory_records(self) -> List[Dict[str, Any]]:
        """获取产品库存流水"""
        data = self.load_data()
        return data.get("product_inventory_records", [])
    
    def get_admin_password(self) -> Optional[str]:
        import os
        data = self.load_data()
        settings = data.get("system_settings") or {}
        salt = settings.get("admin_password_salt")
        pwd_hash = settings.get("admin_password_hash")
        if salt and pwd_hash:
            return None
        try:
            val = os.environ.get("APP_ADMIN_PASSWORD")
            if val:
                return val
        except Exception:
            pass
        try:
            sec = getattr(st, "secrets", None)
            if sec and "admin_password" in sec:
                return sec["admin_password"]
        except Exception:
            pass
        return "admin"
    
    def verify_admin_password(self, password: str) -> bool:
        if not password:
            return False
        data = self.load_data()
        settings = data.get("system_settings") or {}
        salt = settings.get("admin_password_salt")
        pwd_hash = settings.get("admin_password_hash")
        if salt and pwd_hash:
            calc = self._hash_password(str(password), str(salt))
            return secrets.compare_digest(str(calc), str(pwd_hash))
        expected = self.get_admin_password()
        return secrets.compare_digest(str(password), str(expected))
    
    def set_admin_password(self, password: str) -> bool:
        data = self.load_data()
        settings = data.get("system_settings") or {}
        salt = secrets.token_hex(16)
        pwd_hash = self._hash_password(str(password), str(salt))
        settings["admin_password_salt"] = salt
        settings["admin_password_hash"] = pwd_hash
        data["system_settings"] = settings
        return self.save_data(data)
    
    def ensure_raw_material_from_product(self, product_name: str, unit_target: str = UnitType.KG.value) -> Tuple[bool, str]:
        data = self.load_data()
        inventory = data.get(DataCategory.PRODUCT_INVENTORY.value, [])
        materials = data.get(DataCategory.RAW_MATERIALS.value, [])
        records = data.get(DataCategory.INVENTORY_RECORDS.value, [])
        prod = next((p for p in inventory if str(p.get("product_name", "")).strip() == str(product_name).strip()), None)
        if not prod:
            return False, f"未找到成品库存：{product_name}"
        # 检查原材料是否已存在
        mat = next((m for m in materials if str(m.get("name", "")).strip() == str(product_name).strip()), None)
        if not mat:
            new_id = self._get_next_id(materials)
            mat = {
                "id": new_id,
                "name": product_name,
                "supplier": "",
                "spec": "",
                "unit": unit_target,
                "stock_quantity": 0.0,
                "created_date": datetime.now().strftime("%Y-%m-%d"),
                "last_stock_update": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            }
            materials.append(mat)
        # 同步库存：将成品库存数量换算为原材料库存
        qty_ton = float(prod.get("current_stock", 0.0))
        prod_unit = prod.get("unit", UnitType.TON.value)
        final_qty, success = convert_quantity(qty_ton, prod_unit, unit_target)
        if not success and normalize_unit(prod_unit) != normalize_unit(unit_target):
            logger.warning(f"Unit conversion failed when migrating product to raw: {qty_ton} {prod_unit} -> {unit_target}")
            # 默认换算：吨->kg
            if normalize_unit(prod_unit) in ["ton", "吨"] and normalize_unit(unit_target) in ["kg", "千克"]:
                final_qty = qty_ton * 1000.0
            else:
                final_qty = qty_ton
        # 更新原材料库存为换算后的值
        for i, m in enumerate(materials):
            if m.get("name") == product_name:
                materials[i]["stock_quantity"] = final_qty
                materials[i]["last_stock_update"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                break
        # 生成一条台账记录以体现初始同步（adjust_in）
        new_rec_id = self._get_next_id(records)
        records.append({
            "id": new_rec_id,
            "date": datetime.now().strftime("%Y-%m-%d"),
            "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "material_id": mat.get("id"),
            "type": StockMovementType.ADJUST_IN.value,
            "quantity": final_qty,
            "unit": unit_target,
            "reason": f"迁移自成品库存: {product_name}",
            "operator": "System",
            "snapshot_stock": final_qty
        })
        data[DataCategory.RAW_MATERIALS.value] = materials
        data[DataCategory.INVENTORY_RECORDS.value] = records
        self.save_data(data)
        return True, "迁移成功"
    
    def cleanup_migrated_raw_materials(self, names: List[str]) -> Tuple[bool, str]:
        data = self.load_data()
        materials = data.get(DataCategory.RAW_MATERIALS.value, [])
        records = data.get(DataCategory.INVENTORY_RECORDS.value, [])
        name_set = set(str(n).strip() for n in names if n)
        target_ids = [m.get("id") for m in materials if str(m.get("name", "")).strip() in name_set]
        new_materials = [m for m in materials if str(m.get("name", "")).strip() not in name_set]
        new_records = []
        for r in records:
            mid = r.get("material_id")
            reason = str(r.get("reason", "") or "")
            if mid in target_ids and reason.startswith("迁移自成品库存:"):
                continue
            new_records.append(r)
        data[DataCategory.RAW_MATERIALS.value] = new_materials
        data[DataCategory.INVENTORY_RECORDS.value] = new_records
        self.save_data(data)
        return True, f"已清除 {len(target_ids)} 项及相关台账"
    
    def audit_and_fix_product_consumption_mismatch(self) -> List[Dict[str, Any]]:
        data = self.load_data()
        records = data.get(DataCategory.PRODUCT_INVENTORY_RECORDS.value, [])
        issues = data.get(DataCategory.MATERIAL_ISSUES.value, [])
        inventory = data.get(DataCategory.PRODUCT_INVENTORY.value, [])
        issue_map = {}
        for iss in issues:
            code = iss.get("issue_code")
            if code:
                issue_map[code] = iss
        def parse_issue_code(reason):
            s = str(reason or "")
            if "生产领料:" in s:
                part = s.split("生产领料:", 1)[1].strip()
                part = part.split("(", 1)[0].strip()
                return part
            return ""
        def expect_norm(name):
            n = str(name or "").strip()
            if n.startswith("YJSNJ-") or n in ["有碱速凝剂", "碱速凝剂"]:
                return "YJSNJ-有碱速凝剂", "有碱速凝剂"
            if n.startswith("WJSNJ-") or n in ["无碱速凝剂"]:
                return "WJSNJ-无碱速凝剂", "无碱速凝剂"
            return n, ""
        def find_item_idx(nm, tp):
            for i, it in enumerate(inventory):
                if str(it.get("name", "")) == nm and str(it.get("type", "")) == tp:
                    return i
            return -1
        before_snapshot = { (it.get("name"), it.get("type")): float(it.get("stock_quantity", 0.0)) for it in inventory }
        fixed = []
        for i, r in enumerate(records):
            if r.get("type") != StockMovementType.OUT.value:
                continue
            pname = r.get("product_name")
            ptype = r.get("product_type", ProductCategory.OTHER.value)
            reason = r.get("reason", "")
            code = parse_issue_code(reason)
            if not code:
                continue
            iss = issue_map.get(code)
            if not iss:
                continue
            lines = iss.get("lines", [])
            exp_name = ""
            for ln in lines:
                if ln.get("item_type") == "product":
                    exp_name = ln.get("item_name", "")
                    break
            if not exp_name:
                continue
            exp_norm_name, exp_norm_type = expect_norm(exp_name)
            if ptype == ProductCategory.ACCELERATOR.value and exp_norm_type == "有碱速凝剂": # 这里需要确认 ACCELERATOR 是什么
                # 实际上 ProductCategory.ACCELERATOR 是 "速凝剂"，而这里有更细分的 "无碱速凝剂"
                # 由于代码中 logic 比较 specific，暂时保留字符串判断，或者需要扩充 Enum
                pass 
            
            # 这里的逻辑比较特定，暂时保留原字符串判断，或者后续优化
            if ptype == "无碱速凝剂" and exp_norm_type == "有碱速凝剂":
                qty = float(r.get("quantity", 0.0))
                old_idx = find_item_idx(pname, ptype)
                new_idx = find_item_idx(exp_norm_name, exp_norm_type)
                if new_idx == -1:
                    new_id = self._get_next_id(inventory)
                    inventory.append({
                        "id": new_id,
                        "name": exp_norm_name,
                        "type": exp_norm_type,
                        "stock_quantity": 0.0,
                        "unit": UnitType.TON.value,
                        "last_update": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    })
                    new_idx = len(inventory) - 1
                if old_idx >= 0:
                    old_stock = float(inventory[old_idx].get("stock_quantity", 0.0))
                    inventory[old_idx]["stock_quantity"] = old_stock + qty
                    inventory[old_idx]["last_update"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                new_stock = float(inventory[new_idx].get("stock_quantity", 0.0))
                inventory[new_idx]["stock_quantity"] = new_stock - qty
                inventory[new_idx]["last_update"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                r["product_name"] = exp_norm_name
                r["product_type"] = exp_norm_type
                r["snapshot_stock"] = float(inventory[new_idx].get("stock_quantity", 0.0))
                fixed.append({
                    "record_id": r.get("id"),
                    "issue_code": code,
                    "old_name": pname,
                    "old_type": ptype,
                    "new_name": exp_norm_name,
                    "new_type": exp_norm_type,
                    "quantity": qty
                })
        data[DataCategory.PRODUCT_INVENTORY.value] = inventory
        data[DataCategory.PRODUCT_INVENTORY_RECORDS.value] = records
        self.save_data(data)
        after_snapshot = { (it.get("name"), it.get("type")): float(it.get("stock_quantity", 0.0)) for it in inventory }
        report_rows = []
        for f in fixed:
            old_key = (f["old_name"], f["old_type"])
            new_key = (f["new_name"], f["new_type"])
            report_rows.append({
                "记录ID": f["record_id"],
                "领料单号": f["issue_code"],
                "修正数量(吨)": round(f["quantity"], 6),
                "原产品": f["old_name"],
                "原类型": f["old_type"],
                "新产品": f["new_name"],
                "新类型": f["new_type"],
                "原产品修正后库存(吨)": round(after_snapshot.get(old_key, 0.0), 6),
                "新产品修正后库存(吨)": round(after_snapshot.get(new_key, 0.0), 6)
            })
        return report_rows
    
    def normalize_product_aliases(self):
        """
        规范产品名称别名：
        - 将有碱/无碱速凝剂的别名合并为带编码的规范名称
        - 合并库存条目，累计库存数量
        - 级联更新产品库存台账中的产品名称
        Returns:
            list[dict]: 调整摘要
        """
        data = self.load_data()
        inventory = data.get(DataCategory.PRODUCT_INVENTORY.value, [])
        records = data.get(DataCategory.PRODUCT_INVENTORY_RECORDS.value, [])
        
        alias_map = {
            "YJSNJ-有碱速凝剂": ["有碱速凝剂", "碱速凝剂"],
            "WJSNJ-无碱速凝剂": ["无碱速凝剂"]
        }
        
        adjustments = []
        
        for canonical, aliases in alias_map.items():
            group_names = [canonical] + aliases
            group_items = [item for item in inventory if str(item.get("product_name")) in group_names]
            if not group_items:
                adjustments.append({
                    "规范名称": canonical,
                    "合并来源": ",".join(aliases),
                    "原库存合计(吨)": 0.0,
                    "合并后库存(吨)": 0.0,
                    "删除条目数": 0,
                    "更新台账记录数": 0
                })
                continue
            
            total_stock = sum(float(it.get("current_stock", 0.0)) for it in group_items)
            canonical_item = next((it for it in group_items if it.get("product_name") == canonical), None)
            
            if canonical_item is None:
                canonical_item = {
                    "id": self._get_next_id(inventory),
                    "product_name": canonical,
                    "type": "有碱速凝剂" if "有碱速凝剂" in canonical else ("无碱速凝剂" if "无碱速凝剂" in canonical else group_items[0].get("type", ProductCategory.OTHER.value)),
                    "current_stock": 0.0,
                    "unit": group_items[0].get("unit", UnitType.TON.value),
                    "last_update": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                }
                inventory.append(canonical_item)
            
            prev_total = float(canonical_item.get("current_stock", 0.0))
            canonical_item["current_stock"] = total_stock
            canonical_item["last_update"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            desired_type = "有碱速凝剂" if "有碱速凝剂" in canonical else "无碱速凝剂"
            canonical_item["type"] = desired_type
            
            removed_count = 0
            new_inventory = []
            for it in inventory:
                if it.get("product_name") in aliases:
                    removed_count += 1
                    continue
                if it.get("product_name") == canonical:
                    it["type"] = desired_type
                new_inventory.append(it)
            inventory = new_inventory
            
            updated_records = 0
            for r in records:
                if r.get("product_name") in group_names:
                    r["product_name"] = canonical
                    r["product_type"] = desired_type
                    updated_records += 1
            
            adjustments.append({
                "规范名称": canonical,
                "合并来源": ",".join(aliases),
                "原库存合计(吨)": round(prev_total, 4),
                "合并后库存(吨)": round(total_stock, 4),
                "删除条目数": removed_count,
                "更新台账记录数": updated_records
            })
        
        data[DataCategory.PRODUCT_INVENTORY.value] = inventory
        data[DataCategory.PRODUCT_INVENTORY_RECORDS.value] = records
        self.save_data(data)
        return adjustments

    def delete_product_inventory_item(self, product_id: int) -> bool:
        """删除产品库存条目"""
        data = self.load_data()
        inventory = data.get(DataCategory.PRODUCT_INVENTORY.value, [])
        
        new_inventory = [item for item in inventory if item.get("id") != product_id]
        
        if len(new_inventory) < len(inventory):
            data[DataCategory.PRODUCT_INVENTORY.value] = new_inventory
            return self.save_data(data)
        return False
    
    # 混凝土实验
    def get_all_concrete_experiments(self) -> List[Dict[str, Any]]:
        """获取所有混凝土实验"""
        data = self.load_data()
        return data.get(DataCategory.CONCRETE_EXPERIMENTS.value, [])
    
    def add_concrete_experiment(self, experiment_data: Union[Dict[str, Any], ConcreteExperiment]) -> bool:
        """添加新混凝土实验"""
        data = self.load_data()
        experiments = data.get(DataCategory.CONCRETE_EXPERIMENTS.value, [])
        
        # 生成新ID
        new_id = self._get_next_id(experiments)
        
        if isinstance(experiment_data, ConcreteExperiment):
            experiment_data.id = new_id
            if not experiment_data.created_at:
                experiment_data.created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            exp_dict = experiment_data.model_dump(mode='json')
        else:
            experiment_data["id"] = new_id
            if "created_at" not in experiment_data:
                experiment_data["created_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            try:
                ConcreteExperiment(**experiment_data)
            except Exception as e:
                logger.warning(f"ConcreteExperiment validation warning: {e}")
            exp_dict = experiment_data
            
        experiments.append(exp_dict)
        data[DataCategory.CONCRETE_EXPERIMENTS.value] = experiments
        return self.save_data(data)
    
    def delete_concrete_experiment(self, experiment_id: int) -> bool:
        """删除混凝土实验"""
        data = self.load_data()
        experiments = data.get(DataCategory.CONCRETE_EXPERIMENTS.value, [])
        
        new_experiments = [e for e in experiments if e.get("id") != experiment_id]
        
        if len(new_experiments) < len(experiments):
            data[DataCategory.CONCRETE_EXPERIMENTS.value] = new_experiments
            return self.save_data(data)
        return False

    def update_concrete_experiment(self, experiment_id: int, updated_fields: Dict[str, Any]) -> bool:
        """更新混凝土实验"""
        data = self.load_data()
        experiments = data.get(DataCategory.CONCRETE_EXPERIMENTS.value, [])
        updated = False
        for i, exp in enumerate(experiments):
            if exp.get("id") == experiment_id:
                experiments[i].update(updated_fields)
                experiments[i]["last_modified"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                updated = True
                break
        if updated:
            data[DataCategory.CONCRETE_EXPERIMENTS.value] = experiments
            return self.save_data(data)
        return False


    # -------------------- 采购收货管理 --------------------
    def get_all_goods_receipts(self) -> List[Dict[str, Any]]:
        """获取所有采购收货单"""
        data = self.load_data()
        return data.get(DataCategory.GOODS_RECEIPTS.value, [])

    def add_goods_receipt(self, receipt_data: Union[Dict[str, Any], GoodsReceipt]) -> bool:
        """创建采购收货单"""
        data = self.load_data()
        receipts = data.get(DataCategory.GOODS_RECEIPTS.value, [])
        
        if isinstance(receipt_data, GoodsReceipt):
            new_receipt = receipt_data.model_dump(mode='json')
        else:
            new_receipt = receipt_data.copy()
            
        # 生成ID
        new_id = self._get_next_id(receipts)
        new_receipt["id"] = new_id
        
        # 补充默认字段
        if "created_at" not in new_receipt:
            new_receipt["created_at"] = datetime.now().strftime(DATETIME_FORMAT)
        if "status" not in new_receipt:
            new_receipt["status"] = ReceiptStatus.DRAFT.value
        if not new_receipt.get("receipt_code"):
            new_receipt["receipt_code"] = f"GR-{datetime.now().strftime('%Y%m%d')}-{new_id:04d}"
            
        receipts.append(new_receipt)
        data[DataCategory.GOODS_RECEIPTS.value] = receipts
        return self.save_data(data)

    def update_goods_receipt(self, receipt_id: int, updates: Dict[str, Any]) -> bool:
        """更新采购收货单"""
        data = self.load_data()
        receipts = data.get(DataCategory.GOODS_RECEIPTS.value, [])
        
        for i, r in enumerate(receipts):
            if r.get("id") == receipt_id:
                receipts[i].update(updates)
                data[DataCategory.GOODS_RECEIPTS.value] = receipts
                return self.save_data(data)
        return False

    def delete_goods_receipt(self, receipt_id: int) -> bool:
        """删除采购入库单"""
        data = self.load_data()
        receipts = data.get(DataCategory.GOODS_RECEIPTS.value, [])
        new_receipts = [r for r in receipts if r.get("id") != receipt_id]
        if len(new_receipts) < len(receipts):
            data[DataCategory.GOODS_RECEIPTS.value] = new_receipts
            return self.save_data(data)
        return False

    # -------------------- 销售发货管理 --------------------
    def get_all_shipping_orders(self) -> List[Dict[str, Any]]:
        """获取所有销售发货单"""
        data = self.load_data()
        return data.get(DataCategory.SHIPPING_ORDERS.value, [])

    def add_shipping_order(self, order_data: Union[Dict[str, Any], ShippingOrder]) -> bool:
        """创建销售发货单"""
        data = self.load_data()
        orders = data.get(DataCategory.SHIPPING_ORDERS.value, [])
        
        if isinstance(order_data, ShippingOrder):
            new_order = order_data.model_dump(mode='json')
        else:
            new_order = order_data.copy()
            
        # 生成ID
        new_id = self._get_next_id(orders)
        new_order["id"] = new_id
        
        # 补充默认字段
        if "created_at" not in new_order:
            new_order["created_at"] = datetime.now().strftime(DATETIME_FORMAT)
        if "status" not in new_order:
            new_order["status"] = ShippingStatus.DRAFT.value
        if not new_order.get("shipping_code"):
            new_order["shipping_code"] = f"SHIP-{datetime.now().strftime('%Y%m%d')}-{new_id:04d}"
            
        orders.append(new_order)
        data[DataCategory.SHIPPING_ORDERS.value] = orders
        return self.save_data(data)

    def update_shipping_order(self, order_id: int, updates: Dict[str, Any]) -> bool:
        """更新销售发货单"""
        data = self.load_data()
        orders = data.get(DataCategory.SHIPPING_ORDERS.value, [])
        
        for i, o in enumerate(orders):
            if o.get("id") == order_id:
                orders[i].update(updates)
                data[DataCategory.SHIPPING_ORDERS.value] = orders
                return self.save_data(data)
        return False

    def delete_shipping_order(self, order_id: int) -> bool:
        """删除发货单"""
        data = self.load_data()
        orders = data.get(DataCategory.SHIPPING_ORDERS.value, [])
        new_orders = [o for o in orders if o.get("id") != order_id]
        if len(new_orders) < len(orders):
            data[DataCategory.SHIPPING_ORDERS.value] = new_orders
            return self.save_data(data)
        return False

    # -------------------- 数据导出/导入功能 --------------------
    def export_to_excel(self):
        """将所有数据导出到Excel文件"""
        try:
            # 创建内存中的Excel文件
            output = BytesIO()
            
            # 使用pandas的ExcelWriter
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # 获取所有数据
                data = self.load_data()
                
                # 导出各个数据表
                # 1. 项目数据
                if data.get(DataCategory.PROJECTS.value):
                    projects_df = pd.DataFrame(data[DataCategory.PROJECTS.value])
                    projects_df.to_excel(writer, sheet_name='项目', index=False)
                
                # 2. 实验数据
                if data.get(DataCategory.EXPERIMENTS.value):
                    experiments_df = pd.DataFrame(data[DataCategory.EXPERIMENTS.value])
                    experiments_df.to_excel(writer, sheet_name='实验', index=False)
                
                # 3. 原材料数据
                if data.get(DataCategory.RAW_MATERIALS.value):
                    raw_materials_df = pd.DataFrame(data[DataCategory.RAW_MATERIALS.value])
                    raw_materials_df.to_excel(writer, sheet_name='原材料', index=False)
                
                # 4. 合成实验记录
                if data.get(DataCategory.SYNTHESIS_RECORDS.value):
                    synthesis_df = pd.DataFrame(data[DataCategory.SYNTHESIS_RECORDS.value])
                    synthesis_df.to_excel(writer, sheet_name='合成实验', index=False)
                
                # 5. 成品减水剂
                if data.get(DataCategory.PRODUCTS.value):
                    products_df = pd.DataFrame(data[DataCategory.PRODUCTS.value])
                    products_df.to_excel(writer, sheet_name='成品减水剂', index=False)
                
                # 6. 净浆实验
                if data.get(DataCategory.PASTE_EXPERIMENTS.value):
                    paste_df = pd.DataFrame(data[DataCategory.PASTE_EXPERIMENTS.value])
                    paste_df.to_excel(writer, sheet_name='净浆实验', index=False)
                
                # 7. 砂浆实验
                if data.get(DataCategory.MORTAR_EXPERIMENTS.value):
                    mortar_df = pd.DataFrame(data[DataCategory.MORTAR_EXPERIMENTS.value])
                    mortar_df.to_excel(writer, sheet_name='砂浆实验', index=False)
                
                # 8. 混凝土实验
                if data.get(DataCategory.CONCRETE_EXPERIMENTS.value):
                    concrete_df = pd.DataFrame(data[DataCategory.CONCRETE_EXPERIMENTS.value])
                    concrete_df.to_excel(writer, sheet_name='混凝土实验', index=False)
                
                # 9. 性能数据
                if data.get(DataCategory.PERFORMANCE_DATA.value):
                    perf_data = data[DataCategory.PERFORMANCE_DATA.value]
                    if perf_data.get("synthesis"):
                        perf_synth_df = pd.DataFrame(perf_data["synthesis"])
                        perf_synth_df.to_excel(writer, sheet_name='合成性能数据', index=False)
                
                # 10. 数据字典（说明）
                metadata = {
                    'Sheet名称': ['项目', '实验', '原材料', '合成实验', '成品减水剂', '净浆实验', '砂浆实验', '混凝土实验', '合成性能数据'],
                    '描述': [
                        '项目基本信息和管理信息',
                        '实验计划和执行信息',
                        '原材料库管理信息',
                        '合成实验详细记录',
                        '成品减水剂信息',
                        '净浆实验测试数据',
                        '砂浆实验测试数据',
                        '混凝土实验测试数据',
                        '合成性能测试数据'
                    ]
                }
                metadata_df = pd.DataFrame(metadata)
                metadata_df.to_excel(writer, sheet_name='数据字典', index=False)
            
            # 获取二进制数据
            excel_data = output.getvalue()
            b64 = base64.b64encode(excel_data).decode()
            href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="聚羧酸减水剂研发数据_{datetime.now().strftime("%Y%m%d")}.xlsx">点击下载 Excel 文件</a>'
            
            return href
        
        except Exception as e:
            st.error(f"导出数据失败: {e}")
            return None

    def export_experiment_report(self, experiment_type, experiment_id, strength_y_max=None, strength_chart_type="line"):
        try:
            data = self.load_data()
            if experiment_type == "mortar":
                records = data.get(DataCategory.MORTAR_EXPERIMENTS.value, [])
                sheet_title = "砂浆实验报告"
            elif experiment_type == "concrete":
                records = data.get(DataCategory.CONCRETE_EXPERIMENTS.value, [])
                sheet_title = "混凝土实验报告"
            else:
                st.error("不支持的实验类型")
                return None
            target = None
            for r in records:
                if r.get("id") == experiment_id:
                    target = r
                    break
            if not target:
                st.error("未找到指定实验记录")
                return None
            recipes = target.get("test_recipes") or []
            if not recipes:
                st.warning("该实验没有测试配方数据，报告将不包含强度曲线")
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_title
            if experiment_type == "mortar":
                self._fill_mortar_report_sheet(ws, target, strength_y_max, strength_chart_type)
            else:
                self._fill_concrete_report_sheet(ws, target, strength_y_max, strength_chart_type)
            output = BytesIO()
            wb.save(output)
            excel_data = output.getvalue()
            b64 = base64.b64encode(excel_data).decode()
            filename = f"{sheet_title}_{experiment_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="{filename}">下载 {sheet_title}</a>'
            return href
        except Exception as e:
            st.error(f"导出实验报告失败: {e}")
            return None

    def _fill_mortar_report_sheet(self, ws, experiment, strength_y_max=None, strength_chart_type="line"):
        title_font = Font(size=16, bold=True)
        header_font = Font(bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        row = 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        cell = ws.cell(row=row, column=1, value="混凝土外加剂匀质性试验报告（砂浆）")
        cell.font = title_font
        cell.alignment = center_align
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        std_cell = ws.cell(row=row, column=1, value="执行标准：GB/T 8077-2023，产品标准：GB/T 8076-2025")
        std_cell.alignment = center_align
        row += 2
        headers = ["实验ID", "关联配方/外加剂名称", "试验日期", "试验人员", "水灰比", "外加剂掺量(%)", "砂含水率(%)", "备注"]
        values = [
            experiment.get("id"),
            experiment.get("formula_name"),
            experiment.get("test_date"),
            experiment.get("operator"),
            experiment.get("water_cement_ratio"),
            experiment.get("admixture_dosage"),
            experiment.get("sand_moisture"),
            experiment.get("notes"),
        ]
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = header_font
            cell.alignment = center_align
        row += 1
        for col, v in enumerate(values, start=1):
            ws.cell(row=row, column=col, value=v)
        row += 2
        binders = ((experiment.get("materials") or {}).get("binders") or [])
        aggregates = ((experiment.get("materials") or {}).get("aggregates") or [])
        ws.cell(row=row, column=1, value="材料类型").font = header_font
        ws.cell(row=row, column=2, value="材料名称").font = header_font
        ws.cell(row=row, column=3, value="用量(g)").font = header_font
        ws.cell(row=row, column=4, value="备注").font = header_font
        row += 1
        for item in binders:
            ws.cell(row=row, column=1, value="胶凝材料")
            ws.cell(row=row, column=2, value=item.get("name"))
            ws.cell(row=row, column=3, value=item.get("dosage"))
            row += 1
        for item in aggregates:
            ws.cell(row=row, column=1, value="细骨料")
            ws.cell(row=row, column=2, value=item.get("name"))
            ws.cell(row=row, column=3, value=item.get("dosage"))
            row += 1
        row += 2
        recipes = experiment.get("test_recipes") or []
        def age_sort_key(age):
            s = str(age)
            digits = "".join(ch for ch in s if ch.isdigit())
            if digits:
                try:
                    return int(digits)
                except Exception:
                    return 0
            return 0
        ages_set = set()
        for r in recipes:
            perf = r.get("performance") or {}
            cs = perf.get("compressive_strengths") or {}
            for k in cs.keys():
                ages_set.add(k)
        ages = sorted(list(ages_set), key=age_sort_key)
        if ages and recipes:
            ws.cell(row=row, column=1, value="龄期").font = header_font
            col_offset = 2
            for idx, r in enumerate(recipes):
                name = r.get("name") or f"配方{idx + 1}"
                ws.cell(row=row, column=col_offset + idx, value=name).font = header_font
            start_data_row = row + 1
            for r_idx, age in enumerate(ages):
                ws.cell(row=start_data_row + r_idx, column=1, value=str(age))
                for c_idx, r in enumerate(recipes):
                    perf = r.get("performance") or {}
                    cs = perf.get("compressive_strengths") or {}
                    val = cs.get(age)
                    if val is not None:
                        ws.cell(row=start_data_row + r_idx, column=col_offset + c_idx, value=float(val))
            if strength_chart_type == "bar":
                chart = BarChart()
            else:
                chart = LineChart()
            chart.title = "砂浆抗压强度发展曲线"
            chart.y_axis.title = "抗压强度(MPa)"
            chart.x_axis.title = "龄期"
            if strength_y_max and strength_y_max > 0:
                chart.y_axis.scaling.max = float(strength_y_max)
            data_ref = Reference(ws, min_col=2, max_col=1 + len(recipes), min_row=row, max_row=start_data_row + len(ages) - 1)
            chart.add_data(data_ref, titles_from_data=True)
            cats_ref = Reference(ws, min_col=1, min_row=start_data_row, max_row=start_data_row + len(ages) - 1)
            chart.set_categories(cats_ref)
            chart.legend.title = "配方"
            chart_row = start_data_row + len(ages) + 2
            ws.add_chart(chart, f"A{chart_row}")
            ratio_header_row = chart_row + 15
        else:
            ratio_header_row = row + 2
        ws.cell(row=ratio_header_row, column=1, value="指标").font = header_font
        ws.cell(row=ratio_header_row, column=2, value="数值").font = header_font
        ratio_data = [
            ("水灰比", experiment.get("water_cement_ratio")),
            ("外加剂掺量(%)", experiment.get("admixture_dosage")),
            ("砂含水率(%)", experiment.get("sand_moisture")),
        ]
        for idx, (name, val) in enumerate(ratio_data, start=1):
            ws.cell(row=ratio_header_row + idx, column=1, value=name)
            ws.cell(row=ratio_header_row + idx, column=2, value=val)
        bar = BarChart()
        bar.title = "砂浆配合比关键参数"
        bar.y_axis.title = "数值"
        bar.x_axis.title = "参数"
        data_ref = Reference(ws, min_col=2, min_row=ratio_header_row, max_row=ratio_header_row + len(ratio_data))
        bar.add_data(data_ref, titles_from_data=True)
        cats_ref = Reference(ws, min_col=1, min_row=ratio_header_row + 1, max_row=ratio_header_row + len(ratio_data))
        bar.set_categories(cats_ref)
        ws.add_chart(bar, f"E{ratio_header_row}")

    def _fill_concrete_report_sheet(self, ws, experiment, strength_y_max=None, strength_chart_type="line"):
        title_font = Font(size=16, bold=True)
        header_font = Font(bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        row = 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
        cell = ws.cell(row=row, column=1, value="混凝土外加剂性能试验报告（拌和物与强度）")
        cell.font = title_font
        cell.alignment = center_align
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
        std_cell = ws.cell(row=row, column=1, value="执行标准：GB/T 8076-2025，试验方法：GB/T 8077-2023 相关条款")
        std_cell.alignment = center_align
        row += 2
        headers = [
            "实验ID",
            "关联配方",
            "测试日期",
            "操作人",
            "水灰比",
            "砂率(%)",
            "单位用量(kg/m³)",
            "外加剂掺量(%)",
            "砂含水率(%)",
            "石子含水率(%)",
        ]
        values = [
            experiment.get("id"),
            experiment.get("formula_name"),
            experiment.get("test_date"),
            experiment.get("operator"),
            experiment.get("water_cement_ratio"),
            experiment.get("sand_ratio"),
            experiment.get("unit_weight"),
            experiment.get("admixture_dosage"),
            experiment.get("sand_moisture"),
            experiment.get("stone_moisture"),
        ]
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = header_font
            cell.alignment = center_align
        row += 1
        for col, v in enumerate(values, start=1):
            ws.cell(row=row, column=col, value=v)
        row += 2
        binders = ((experiment.get("materials") or {}).get("binders") or [])
        aggregates = ((experiment.get("materials") or {}).get("aggregates") or [])
        ws.cell(row=row, column=1, value="材料类型").font = header_font
        ws.cell(row=row, column=2, value="材料名称").font = header_font
        ws.cell(row=row, column=3, value="用量(kg/m³)").font = header_font
        row += 1
        for item in binders:
            ws.cell(row=row, column=1, value="胶凝材料")
            ws.cell(row=row, column=2, value=item.get("name"))
            ws.cell(row=row, column=3, value=item.get("dosage"))
            row += 1
        for item in aggregates:
            ws.cell(row=row, column=1, value="骨料")
            ws.cell(row=row, column=2, value=item.get("name"))
            ws.cell(row=row, column=3, value=item.get("dosage"))
            row += 1
        row += 2
        recipes = experiment.get("test_recipes") or []
        def age_sort_key(age):
            s = str(age)
            digits = "".join(ch for ch in s if ch.isdigit())
            if digits:
                try:
                    return int(digits)
                except Exception:
                    return 0
            return 0
        ages_set = set()
        for r in recipes:
            perf = r.get("performance") or {}
            strengths = perf.get("strengths") or {}
            for k in strengths.keys():
                ages_set.add(k)
        ages = sorted(list(ages_set), key=age_sort_key)
        if ages and recipes:
            ws.cell(row=row, column=1, value="龄期").font = header_font
            col_offset = 2
            for idx, r in enumerate(recipes):
                name = r.get("name") or f"配方{idx + 1}"
                ws.cell(row=row, column=col_offset + idx, value=name).font = header_font
            start_data_row = row + 1
            for r_idx, age in enumerate(ages):
                ws.cell(row=start_data_row + r_idx, column=1, value=str(age))
                for c_idx, r in enumerate(recipes):
                    perf = r.get("performance") or {}
                    strengths = perf.get("strengths") or {}
                    val = strengths.get(age)
                    if val is not None:
                        ws.cell(row=start_data_row + r_idx, column=col_offset + c_idx, value=float(val))
            if strength_chart_type == "bar":
                chart = BarChart()
            else:
                chart = LineChart()
            chart.title = "混凝土抗压强度发展曲线"
            chart.y_axis.title = "抗压强度(MPa)"
            chart.x_axis.title = "龄期"
            if strength_y_max and strength_y_max > 0:
                chart.y_axis.scaling.max = float(strength_y_max)
            data_ref = Reference(ws, min_col=2, max_col=1 + len(recipes), min_row=row, max_row=start_data_row + len(ages) - 1)
            chart.add_data(data_ref, titles_from_data=True)
            cats_ref = Reference(ws, min_col=1, min_row=start_data_row, max_row=start_data_row + len(ages) - 1)
            chart.set_categories(cats_ref)
            chart.legend.title = "配方"
            chart_row = start_data_row + len(ages) + 2
            ws.add_chart(chart, f"A{chart_row}")
            ratio_header_row = chart_row + 15
        else:
            ratio_header_row = row + 2
        ws.cell(row=ratio_header_row, column=1, value="指标").font = header_font
        ws.cell(row=ratio_header_row, column=2, value="数值").font = header_font
        ratio_data = [
            ("水灰比", experiment.get("water_cement_ratio")),
            ("砂率(%)", experiment.get("sand_ratio")),
            ("单位用量(kg/m³)", experiment.get("unit_weight")),
            ("外加剂掺量(%)", experiment.get("admixture_dosage")),
        ]
        for idx, (name, val) in enumerate(ratio_data, start=1):
            ws.cell(row=ratio_header_row + idx, column=1, value=name)
            ws.cell(row=ratio_header_row + idx, column=2, value=val)
        bar = BarChart()
        bar.title = "混凝土配合比关键参数"
        bar.y_axis.title = "数值"
        bar.x_axis.title = "参数"
        data_ref = Reference(ws, min_col=2, min_row=ratio_header_row, max_row=ratio_header_row + len(ratio_data))
        bar.add_data(data_ref, titles_from_data=True)
        cats_ref = Reference(ws, min_col=1, min_row=ratio_header_row + 1, max_row=ratio_header_row + len(ratio_data))
        bar.set_categories(cats_ref)
        ws.add_chart(bar, f"E{ratio_header_row}")

    def import_from_excel(self, uploaded_file):
        """从Excel文件导入数据"""
        try:
            # 读取Excel文件
            excel_file = pd.ExcelFile(uploaded_file)
            
            # 获取现有数据
            data = self.load_data()
            
            # 导入各个工作表
            import_summary = []
            
            # 1. 导入项目数据
            if '项目' in excel_file.sheet_names:
                projects_df = pd.read_excel(excel_file, sheet_name='项目')
                if not projects_df.empty:
                    # 转换为字典列表
                    projects_list = projects_df.to_dict('records')
                    # 确保ID字段存在
                    for i, project in enumerate(projects_list, 1):
                        if 'id' not in project or pd.isna(project['id']):
                            project['id'] = i
                    data['projects'] = projects_list
                    import_summary.append(f"项目: {len(projects_list)} 条")
            
            # 2. 导入实验数据
            if '实验' in excel_file.sheet_names:
                experiments_df = pd.read_excel(excel_file, sheet_name='实验')
                if not experiments_df.empty:
                    experiments_list = experiments_df.to_dict('records')
                    for i, exp in enumerate(experiments_list, 1):
                        if 'id' not in exp or pd.isna(exp['id']):
                            exp['id'] = i
                    data['experiments'] = experiments_list
                    import_summary.append(f"实验: {len(experiments_list)} 条")
            
            # 3. 导入原材料数据
            if '原材料' in excel_file.sheet_names:
                materials_df = pd.read_excel(excel_file, sheet_name='原材料')
                if not materials_df.empty:
                    materials_list = materials_df.to_dict('records')
                    for i, mat in enumerate(materials_list, 1):
                        if 'id' not in mat or pd.isna(mat['id']):
                            mat['id'] = i
                    data['raw_materials'] = materials_list
                    import_summary.append(f"原材料: {len(materials_list)} 条")
            
            # 4. 导入合成实验数据
            if '合成实验' in excel_file.sheet_names:
                synthesis_df = pd.read_excel(excel_file, sheet_name='合成实验')
                if not synthesis_df.empty:
                    synthesis_list = synthesis_df.to_dict('records')
                    for i, record in enumerate(synthesis_list, 1):
                        if 'id' not in record or pd.isna(record['id']):
                            record['id'] = i
                    data['synthesis_records'] = synthesis_list
                    import_summary.append(f"合成实验: {len(synthesis_list)} 条")
            
            # 5. 导入成品减水剂数据
            if '成品减水剂' in excel_file.sheet_names:
                products_df = pd.read_excel(excel_file, sheet_name='成品减水剂')
                if not products_df.empty:
                    products_list = products_df.to_dict('records')
                    for i, product in enumerate(products_list, 1):
                        if 'id' not in product or pd.isna(product['id']):
                            product['id'] = i
                    data['products'] = products_list
                    import_summary.append(f"成品减水剂: {len(products_list)} 条")
            
            # 6. 导入净浆实验数据
            if '净浆实验' in excel_file.sheet_names:
                paste_df = pd.read_excel(excel_file, sheet_name='净浆实验')
                if not paste_df.empty:
                    paste_list = paste_df.to_dict('records')
                    for i, exp in enumerate(paste_list, 1):
                        if 'id' not in exp or pd.isna(exp['id']):
                            exp['id'] = i
                    data['paste_experiments'] = paste_list
                    import_summary.append(f"净浆实验: {len(paste_list)} 条")
            
            # 7. 导入砂浆实验数据
            if '砂浆实验' in excel_file.sheet_names:
                mortar_df = pd.read_excel(excel_file, sheet_name='砂浆实验')
                if not mortar_df.empty:
                    mortar_list = mortar_df.to_dict('records')
                    for i, exp in enumerate(mortar_list, 1):
                        if 'id' not in exp or pd.isna(exp['id']):
                            exp['id'] = i
                    data['mortar_experiments'] = mortar_list
                    import_summary.append(f"砂浆实验: {len(mortar_list)} 条")
            
            # 8. 导入混凝土实验数据
            if '混凝土实验' in excel_file.sheet_names:
                concrete_df = pd.read_excel(excel_file, sheet_name='混凝土实验')
                if not concrete_df.empty:
                    concrete_list = concrete_df.to_dict('records')
                    for i, exp in enumerate(concrete_list, 1):
                        if 'id' not in exp or pd.isna(exp['id']):
                            exp['id'] = i
                    data['concrete_experiments'] = concrete_list
                    import_summary.append(f"混凝土实验: {len(concrete_list)} 条")
            
            # 9. 导入合成性能数据
            if '合成性能数据' in excel_file.sheet_names:
                perf_df = pd.read_excel(excel_file, sheet_name='合成性能数据')
                if not perf_df.empty:
                    perf_list = perf_df.to_dict('records')
                    if 'performance_data' not in data:
                        data['performance_data'] = {}
                    data['performance_data']['synthesis'] = perf_list
                    import_summary.append(f"合成性能数据: {len(perf_list)} 条")
            
            # 保存导入的数据
            if self.save_data(data):
                return True, "，".join(import_summary)
            else:
                return False, "保存导入数据失败"
                
        except Exception as e:
            return False, f"导入数据失败: {str(e)}"

    def get_json_content(self):
        """获取当前数据的JSON字符串"""
        if self.data_file.exists():
            with open(self.data_file, 'r', encoding='utf-8') as f:
                return f.read()
        return "{}"

    def import_from_json(self, json_content):
        """从JSON字符串导入数据（完全覆盖）"""
        try:
            data = json.loads(json_content)
            # 验证结构
            if not isinstance(data, dict):
                return False, "数据格式错误：根节点必须是对象"
            
            # 确保必要字段存在
            self._ensure_data_structure(data)
            
            # 保存
            if self.save_data(data):
                return True, "数据恢复成功"
            else:
                return False, "保存数据失败"
        except json.JSONDecodeError:
            return False, "JSON格式解析失败"
        except Exception as e:
            return False, f"导入失败: {str(e)}"

